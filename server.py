import base64
import hashlib
import json
import os
import re
import sqlite3
import uuid
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote

from flask import Flask, Response, g, has_request_context, jsonify, request, session

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - runtime safeguard when dependency is missing
    load_workbook = None

BASE_DIR = Path(__file__).resolve().parent
BASE_DB_PATH = Path(os.getenv("PLANNER_DB_PATH", str(BASE_DIR / "planner.db")))
if not BASE_DB_PATH.is_absolute():
    BASE_DB_PATH = (BASE_DIR / BASE_DB_PATH).resolve()

DB_VARIANTS_DIR = Path(os.getenv("PLANNER_DB_VARIANTS_DIR", str(BASE_DB_PATH.parent / "variants")))
if not DB_VARIANTS_DIR.is_absolute():
    DB_VARIANTS_DIR = (BASE_DIR / DB_VARIANTS_DIR).resolve()

ACTIVE_DB_KEY_PATH = Path(os.getenv("PLANNER_ACTIVE_DB_FILE", str(BASE_DB_PATH.parent / ".active_db_key")))
if not ACTIVE_DB_KEY_PATH.is_absolute():
    ACTIVE_DB_KEY_PATH = (BASE_DIR / ACTIVE_DB_KEY_PATH).resolve()

DATABASE_ACCESS_MAP_PATH = Path(
    os.getenv("PLANNER_DB_ACCESS_MAP_FILE", str(BASE_DB_PATH.parent / "database_access_map.json"))
)
if not DATABASE_ACCESS_MAP_PATH.is_absolute():
    DATABASE_ACCESS_MAP_PATH = (BASE_DIR / DATABASE_ACCESS_MAP_PATH).resolve()

DEFAULT_DB_KEY = "default"

app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path="")
app.config["SECRET_KEY"] = os.getenv("APP_SECRET_KEY", "dev-only-change-me")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = str(os.getenv("SESSION_COOKIE_SECURE", "false")).strip().lower() in (
    "1",
    "true",
    "yes",
    "on",
)

STATIONS = [
    {"id": "M1", "name": "Pila optymalizacyjna", "department": "Maszynownia"},
    {"id": "M2", "name": "Strugarka 4-stronna", "department": "Maszynownia"},
    {"id": "M3", "name": "CNC 1", "department": "Maszynownia"},
    {"id": "M4", "name": "CNC 2", "department": "Maszynownia"},
    {"id": "M5", "name": "Czopiarka", "department": "Maszynownia"},
    {"id": "M6", "name": "Frezarka", "department": "Maszynownia"},
    {"id": "M7", "name": "Szlifowanie ram", "department": "Maszynownia"},
    {"id": "M8", "name": "Szlifowanie skrzydel", "department": "Maszynownia"},
    {"id": "M9", "name": "Prasa klejowa", "department": "Maszynownia"},
    {"id": "M10", "name": "Kontrola wymiarowa", "department": "Maszynownia"},
    {"id": "P1", "name": "Przygotowanie lakieru", "department": "Lakiernia"},
    {"id": "P2", "name": "Kabina natryskowa 1", "department": "Lakiernia"},
    {"id": "P3", "name": "Kabina natryskowa 2", "department": "Lakiernia"},
    {"id": "P4", "name": "Suszarnia 1", "department": "Lakiernia"},
    {"id": "P5", "name": "Suszarnia 2", "department": "Lakiernia"},
    {"id": "P6", "name": "Szlif miedzywarstwowy", "department": "Lakiernia"},
    {"id": "P7", "name": "Kontrola powloki", "department": "Lakiernia"},
    {"id": "K1", "name": "Szklenie", "department": "Kompletacja"},
    {"id": "K2", "name": "Okuwanie", "department": "Kompletacja"},
    {"id": "K3", "name": "Montaz uszczelek", "department": "Kompletacja"},
    {"id": "K4", "name": "Montaz akcesoriow", "department": "Kompletacja"},
    {"id": "K5", "name": "Regulacja skrzydel", "department": "Kompletacja"},
    {"id": "K6", "name": "Pakowanie", "department": "Kompletacja"},
    {"id": "K7", "name": "Etykietowanie", "department": "Kompletacja"},
    {"id": "K8", "name": "Kontrola koncowa", "department": "Kompletacja"},
]

MATERIAL_KEYS = ["wood", "corpus", "glass", "hardware", "accessories"]
DEPARTMENTS = ["Maszynownia", "Lakiernia", "Kompletacja", "Kosmetyka"]
USER_VISIBLE_SECTIONS = [
    "orders",
    "kpi",
    "gantt",
    "reports",
    "execution",
    "skills",
    "feedback",
    "archive",
    "users",
    "settings",
]
DEFAULT_MATERIAL_RULES = {
    "Maszynownia": ["wood", "corpus"],
    "Lakiernia": ["wood", "corpus"],
    "Kompletacja": ["glass", "hardware", "accessories"],
    "Kosmetyka": ["accessories"],
}

DEFAULT_TECHNOLOGIES = {
    "IV68 Drewno": {
        "machiningPreferred": ["M1", "M3", "M5", "M7"],
        "paintingPreferred": ["P2", "P4", "P6"],
        "assemblyPreferred": ["K1", "K3", "K6", "K8"],
    },
    "IV78 Drewno": {
        "machiningPreferred": ["M2", "M4", "M6", "M9"],
        "paintingPreferred": ["P2", "P3", "P5"],
        "assemblyPreferred": ["K2", "K4", "K5", "K8"],
    },
    "IV92 Drewno": {
        "machiningPreferred": ["M2", "M4", "M6", "M9", "M10"],
        "paintingPreferred": ["P3", "P5", "P7"],
        "assemblyPreferred": ["K2", "K4", "K5", "K8"],
    },
    "IV68 Drewno-Alu": {
        "machiningPreferred": ["M1", "M3", "M5", "M8"],
        "paintingPreferred": ["P1", "P2", "P4"],
        "assemblyPreferred": ["K1", "K2", "K4", "K8"],
    },
    "IV78 Drewno-Alu": {
        "machiningPreferred": ["M2", "M3", "M6", "M9"],
        "paintingPreferred": ["P1", "P3", "P5"],
        "assemblyPreferred": ["K2", "K3", "K4", "K8"],
    },
    "IV92 Drewno-Alu": {
        "machiningPreferred": ["M2", "M4", "M6", "M9", "M10"],
        "paintingPreferred": ["P3", "P5", "P7"],
        "assemblyPreferred": ["K2", "K4", "K5", "K8"],
    },
    "Drzwi Drewno": {
        "machiningPreferred": ["M1", "M5", "M6", "M8", "M10"],
        "paintingPreferred": ["P2", "P4", "P7"],
        "assemblyPreferred": ["K2", "K4", "K6", "K8"],
    },
    "Drzwi Drewno-Alu": {
        "machiningPreferred": ["M1", "M4", "M6", "M9", "M10"],
        "paintingPreferred": ["P1", "P3", "P5", "P7"],
        "assemblyPreferred": ["K2", "K4", "K5", "K8"],
    },
    "HS": {
        "machiningPreferred": ["M2", "M3", "M4", "M9", "M10"],
        "paintingPreferred": ["P3", "P5", "P7"],
        "assemblyPreferred": ["K1", "K2", "K5", "K6", "K8"],
    },
    "Otwierane na zewnatrz": {
        "machiningPreferred": ["M1", "M3", "M5", "M7", "M10"],
        "paintingPreferred": ["P2", "P4", "P6"],
        "assemblyPreferred": ["K1", "K3", "K5", "K8"],
    },
}
try:
    _raw_attachment_limit = int(os.getenv("MAX_ATTACHMENT_BYTES", str(15 * 1024 * 1024)))
except (TypeError, ValueError):
    _raw_attachment_limit = 15 * 1024 * 1024
MAX_ATTACHMENT_BYTES = max(1024, min(100 * 1024 * 1024, _raw_attachment_limit))
DEFAULT_WEEKDAY_SHIFTS = {"0": 0, "1": 2, "2": 2, "3": 2, "4": 2, "5": 2, "6": 0}


def sanitize_variant_slug(value):
    cleaned = re.sub(r"[^a-z0-9_-]+", "-", str(value or "").strip().lower()).strip("-_")
    return cleaned


def database_key_to_path(key):
    normalized = str(key or "").strip()
    if normalized == DEFAULT_DB_KEY:
        return BASE_DB_PATH
    if normalized.startswith("variant:"):
        slug = sanitize_variant_slug(normalized.split(":", 1)[1])
        if not slug:
            return None
        return DB_VARIANTS_DIR / f"{slug}.db"
    return None


def ensure_database_storage():
    BASE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    DB_VARIANTS_DIR.mkdir(parents=True, exist_ok=True)
    ACTIVE_DB_KEY_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATABASE_ACCESS_MAP_PATH.parent.mkdir(parents=True, exist_ok=True)


def get_active_database_key():
    ensure_database_storage()
    if ACTIVE_DB_KEY_PATH.exists():
        raw = ACTIVE_DB_KEY_PATH.read_text(encoding="utf-8").strip()
        path = database_key_to_path(raw)
        if path and path.exists():
            return raw
    return DEFAULT_DB_KEY


def set_active_database_key(key):
    ensure_database_storage()
    path = database_key_to_path(key)
    if not path or not path.exists():
        raise ValueError("Nie znaleziono wskazanej bazy danych.")
    ACTIVE_DB_KEY_PATH.write_text(str(key), encoding="utf-8")


def get_request_database_key():
    session_key = ""
    if has_request_context():
        session_key = str(session.get("db_key", "")).strip()
    session_path = database_key_to_path(session_key)
    if session_key and session_path and session_path.exists():
        return session_key
    return get_active_database_key()


def set_request_database_key(key):
    if not has_request_context():
        return
    normalized = str(key or "").strip()
    if not normalized:
        session.pop("db_key", None)
        return
    path = database_key_to_path(normalized)
    if not path or not path.exists():
        raise ValueError("Nie znaleziono wskazanej bazy danych.")
    session["db_key"] = normalized


def get_current_database_path():
    key = get_request_database_key()
    path = database_key_to_path(key)
    if not path:
        return BASE_DB_PATH
    return path


def list_database_catalog(active_key=None):
    ensure_database_storage()
    resolved_active_key = str(active_key or "").strip() or get_request_database_key()
    output = [
        {
            "key": DEFAULT_DB_KEY,
            "name": "Domyslna baza",
            "fileName": BASE_DB_PATH.name,
            "active": resolved_active_key == DEFAULT_DB_KEY,
            "variant": False,
        }
    ]
    for file_path in sorted(DB_VARIANTS_DIR.glob("*.db"), key=lambda item: item.name.lower()):
        key = f"variant:{file_path.stem}"
        output.append(
            {
                "key": key,
                "name": file_path.stem,
                "fileName": file_path.name,
                "active": resolved_active_key == key,
                "variant": True,
            }
        )
    return output


def list_database_keys():
    ensure_database_storage()
    keys = [DEFAULT_DB_KEY]
    for file_path in sorted(DB_VARIANTS_DIR.glob("*.db"), key=lambda item: item.name.lower()):
        keys.append(f"variant:{file_path.stem}")
    return keys


def normalize_login_key(value):
    return str(value or "").strip().lower()


def normalize_database_access_map(raw_map, known_keys=None):
    allowed_keys = set(known_keys or list_database_keys())
    output = {}
    if not isinstance(raw_map, dict):
        return output

    for raw_login, raw_keys in raw_map.items():
        login_key = normalize_login_key(raw_login)
        if not login_key:
            continue
        keys_source = raw_keys if isinstance(raw_keys, list) else []
        valid_keys = []
        for raw_key in keys_source:
            key = str(raw_key or "").strip()
            if not key or key not in allowed_keys:
                continue
            if key in valid_keys:
                continue
            valid_keys.append(key)
        output[login_key] = valid_keys
    return output


def load_database_access_map():
    ensure_database_storage()
    if not DATABASE_ACCESS_MAP_PATH.exists():
        return {}
    try:
        parsed = json.loads(DATABASE_ACCESS_MAP_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return normalize_database_access_map(parsed, list_database_keys())


def save_database_access_map(raw_map):
    ensure_database_storage()
    normalized = normalize_database_access_map(raw_map, list_database_keys())
    DATABASE_ACCESS_MAP_PATH.write_text(
        json.dumps(normalized, ensure_ascii=True, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return normalized


def can_user_access_database(user, database_key, access_map=None):
    key = str(database_key or "").strip() or DEFAULT_DB_KEY
    role = str((user or {}).get("role", "")).strip().lower()
    if role == "admin":
        return True
    login_key = normalize_login_key((user or {}).get("login"))
    if not login_key:
        return False
    resolved_access_map = access_map if isinstance(access_map, dict) else load_database_access_map()
    allowed = resolved_access_map.get(login_key)
    if allowed is None:
        return True
    return key in allowed


def list_database_catalog_for_user(user, active_key=None, access_map=None):
    catalog = list_database_catalog(active_key)
    role = str((user or {}).get("role", "")).strip().lower()
    if role == "admin":
        return catalog
    login_key = normalize_login_key((user or {}).get("login"))
    if not login_key:
        return []
    resolved_access_map = access_map if isinstance(access_map, dict) else load_database_access_map()
    allowed = resolved_access_map.get(login_key)
    if allowed is None:
        return catalog
    allowed_set = set(allowed)
    return [item for item in catalog if str(item.get("key") or "") in allowed_set]


def rename_database_access_login(old_login, new_login):
    old_key = normalize_login_key(old_login)
    new_key = normalize_login_key(new_login)
    if not old_key or not new_key or old_key == new_key:
        return
    access_map = load_database_access_map()
    if old_key not in access_map:
        return
    old_values = list(access_map.pop(old_key))
    merged = []
    for key in list(access_map.get(new_key, [])) + old_values:
        key_norm = str(key or "").strip()
        if not key_norm or key_norm in merged:
            continue
        merged.append(key_norm)
    access_map[new_key] = merged
    save_database_access_map(access_map)


def remove_database_access_login(login):
    login_key = normalize_login_key(login)
    if not login_key:
        return
    access_map = load_database_access_map()
    if login_key not in access_map:
        return
    access_map.pop(login_key, None)
    save_database_access_map(access_map)


def db_connect(path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def db():
    return db_connect(get_current_database_path())


def now_iso():
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def normalize_order_status(value):
    status = str(value or "").strip() or "Dokumentacja"
    if status == "Planowanie":
        return "Dokumentacja"
    return status


def normalize_department_status(value):
    status = str(value or "").strip() or "Dokumentacja"
    if status == "Planowanie":
        return "Dokumentacja"
    return status


def normalize_workflow_status(value, fallback="pending"):
    status = str(value or "").strip().lower()
    if status in ("pending", "in_progress", "done"):
        return status
    return fallback


def parse_progress_percent(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    return round(max(0.0, min(100.0, parsed)), 1)


def normalize_position_progress(position_status, progress_value):
    status = normalize_workflow_status(position_status, "pending")
    parsed = parse_progress_percent(progress_value)
    if status == "done":
        return 100.0
    if parsed is None:
        return 1.0 if status == "in_progress" else 0.0
    if status == "in_progress" and parsed <= 0:
        return 1.0
    return round(max(0.0, min(100.0, parsed)), 1)


def normalize_position_department_for_status(position_status, department_status):
    raw_status = str(position_status or "").strip().lower()
    if raw_status == "done":
        return "Zakonczone"
    return normalize_department_status(department_status)


def derive_order_status_from_position_departments(department_statuses):
    normalized = [normalize_department_status(value) for value in department_statuses]
    unique_statuses = set(normalized)
    if unique_statuses == {"Zakonczone"}:
        return "Zakonczone"
    if unique_statuses == {"Dokumentacja"}:
        return "Dokumentacja"
    if unique_statuses.issubset({"Kosmetyka", "Zakonczone"}) and "Kosmetyka" in unique_statuses:
        return "Kosmetyka"
    return "Produkcja"


def sync_order_status_from_positions(connection, order_id):
    order_row = connection.execute(
        "SELECT id, order_status, completed_at FROM orders WHERE id = ?",
        (order_id,),
    ).fetchone()
    if not order_row:
        return False

    # Ujednolicenie historycznych danych: pozycja done powinna byc zakonczona w dziale.
    connection.execute(
        """
        UPDATE order_positions
        SET current_department_status = 'Zakonczone'
        WHERE order_id = ? AND status = 'done' AND COALESCE(current_department_status, '') <> 'Zakonczone'
        """,
        (order_id,),
    )

    status_rows = connection.execute(
        "SELECT current_department_status, status FROM order_positions WHERE order_id = ?",
        (order_id,),
    ).fetchall()
    current_status = normalize_order_status(order_row["order_status"])
    if status_rows:
        next_status = derive_order_status_from_position_departments(
            [
                normalize_position_department_for_status(row["status"], row["current_department_status"])
                for row in status_rows
            ]
        )
    else:
        next_status = current_status

    current_completed_at = normalize_optional_date(order_row["completed_at"])
    if next_status == "Zakonczone":
        next_completed_at = current_completed_at or datetime.now().date().isoformat()
    else:
        next_completed_at = None

    if current_status == next_status and current_completed_at == next_completed_at:
        return False

    connection.execute(
        "UPDATE orders SET order_status = ?, completed_at = ? WHERE id = ?",
        (next_status, next_completed_at, order_id),
    )
    return True


def sync_order_statuses_for_orders(connection, order_ids):
    changed = False
    seen = set()
    for raw_order_id in order_ids:
        order_id = str(raw_order_id or "").strip()
        if not order_id or order_id in seen:
            continue
        seen.add(order_id)
        if sync_order_status_from_positions(connection, order_id):
            changed = True
    return changed


def sync_all_order_statuses(connection):
    order_ids = [row["id"] for row in connection.execute("SELECT id FROM orders").fetchall()]
    return sync_order_statuses_for_orders(connection, order_ids)


def apply_feedback_update_to_position(
    connection,
    row,
    workflow_status=None,
    department_status=None,
    progress_percent=None,
    actor="Brygadzista",
):
    existing_status = normalize_workflow_status(row["status"], "pending")
    existing_department = normalize_department_status(row["current_department_status"])
    existing_progress = normalize_position_progress(existing_status, row["progress_percent"])

    target_status = existing_status if workflow_status is None else normalize_workflow_status(workflow_status, existing_status)
    target_department = existing_department if department_status is None else normalize_department_status(department_status)
    target_progress = existing_progress if progress_percent is None else round(max(0.0, min(100.0, float(progress_percent))), 1)
    started_at = row["started_at"]
    finished_at = row["finished_at"]
    timestamp = now_iso()

    if target_status == "done":
        target_department = "Zakonczone"
        target_progress = 100.0
        started_at = started_at or timestamp
        finished_at = timestamp
    elif target_status == "in_progress":
        if target_department == "Dokumentacja":
            target_department = "Maszynownia"
        started_at = started_at or timestamp
        finished_at = None
        if target_progress <= 0:
            target_progress = 1.0
    else:
        started_at = None
        finished_at = None
        if progress_percent is None:
            target_progress = 0.0

    target_progress = round(max(0.0, min(100.0, target_progress)), 1)

    changed = (
        existing_status != target_status
        or existing_department != target_department
        or abs(existing_progress - target_progress) > 0.0001
        or normalize_optional_text(started_at) != normalize_optional_text(row["started_at"])
        or normalize_optional_text(finished_at) != normalize_optional_text(row["finished_at"])
    )
    if not changed:
        return False

    connection.execute(
        """
        UPDATE order_positions
        SET status = ?, current_department_status = ?, progress_percent = ?, started_at = ?, finished_at = ?
        WHERE id = ?
        """,
        (target_status, target_department, target_progress, started_at, finished_at, row["id"]),
    )

    if actor:
        connection.execute(
            "INSERT INTO feedback_events (id, order_id, position_id, action, actor, at) VALUES (?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), row["order_id"], row["id"], "update", actor, timestamp),
        )
    return True


def column_exists(connection, table_name, column_name):
    rows = connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(row["name"] == column_name for row in rows)


def ensure_column(connection, table_name, column_name, definition_sql):
    if not column_exists(connection, table_name, column_name):
        connection.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition_sql}")


def make_distribution(preferred_ids, all_ids):
    if not all_ids:
        return {}
    preferred = [item for item in preferred_ids if item in all_ids]
    if not preferred:
        equal = round(100 / len(all_ids), 4)
        return {station_id: equal for station_id in all_ids}

    preferred_share = 60.0
    common_share = 40.0
    per_preferred = preferred_share / len(preferred)
    per_all = common_share / len(all_ids)
    allocation = {station_id: per_all for station_id in all_ids}
    for station_id in preferred:
        allocation[station_id] += per_preferred

    total = sum(allocation.values()) or 1
    normalized = {station_id: round((value / total) * 100, 4) for station_id, value in allocation.items()}
    correction = round(100 - sum(normalized.values()), 4)
    first_key = next(iter(normalized.keys()))
    normalized[first_key] = round(normalized[first_key] + correction, 4)
    return normalized


def default_technology_allocations():
    machining_ids = [item["id"] for item in STATIONS if item["department"] == "Maszynownia"]
    painting_ids = [item["id"] for item in STATIONS if item["department"] == "Lakiernia"]
    assembly_ids = [item["id"] for item in STATIONS if item["department"] == "Kompletacja"]
    output = {}
    for tech_name, rules in DEFAULT_TECHNOLOGIES.items():
        output[tech_name] = {
            "machining": make_distribution(rules["machiningPreferred"], machining_ids),
            "painting": make_distribution(rules["paintingPreferred"], painting_ids),
            "assembly": make_distribution(rules["assemblyPreferred"], assembly_ids),
        }
    return output


def initialize(target_db_path=None):
    connection = db_connect(target_db_path or get_current_database_path())

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
          id INTEGER PRIMARY KEY CHECK (id = 1),
          minutes_per_shift INTEGER NOT NULL
        )
        """
    )
    ensure_column(connection, "settings", "minutes_per_shift", "INTEGER NOT NULL DEFAULT 480")
    ensure_column(connection, "settings", "working_days_json", "TEXT NOT NULL DEFAULT '[1,2,3,4,5]'")
    ensure_column(connection, "settings", "calendar_overrides_json", "TEXT NOT NULL DEFAULT '{}'")
    ensure_column(
        connection,
        "settings",
        "weekday_shifts_json",
        "TEXT NOT NULL DEFAULT '{\"0\":0,\"1\":2,\"2\":2,\"3\":2,\"4\":2,\"5\":2,\"6\":0}'",
    )
    ensure_column(connection, "settings", "station_overtime_json", "TEXT NOT NULL DEFAULT '{}'")
    if column_exists(connection, "settings", "work_hours_per_shift"):
        connection.execute(
            """
            UPDATE settings
            SET minutes_per_shift = CASE
              WHEN minutes_per_shift IS NULL OR minutes_per_shift <= 0 THEN work_hours_per_shift * 60
              ELSE minutes_per_shift
            END
            """
        )
    connection.execute(
        "UPDATE settings SET working_days_json = COALESCE(NULLIF(working_days_json, ''), '[1,2,3,4,5]') WHERE id = 1"
    )
    connection.execute(
        "UPDATE settings SET calendar_overrides_json = COALESCE(NULLIF(calendar_overrides_json, ''), '{}') WHERE id = 1"
    )
    connection.execute(
        """
        UPDATE settings
        SET weekday_shifts_json = COALESCE(NULLIF(weekday_shifts_json, ''), '{"0":0,"1":2,"2":2,"3":2,"4":2,"5":2,"6":0}')
        WHERE id = 1
        """
    )
    connection.execute(
        "UPDATE settings SET station_overtime_json = COALESCE(NULLIF(station_overtime_json, ''), '{}') WHERE id = 1"
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
          id TEXT PRIMARY KEY,
          name TEXT NOT NULL,
          department TEXT NOT NULL
        )
        """
    )
    ensure_column(connection, "users", "login", "TEXT")
    ensure_column(connection, "users", "password_hash", "TEXT")
    ensure_column(connection, "users", "role", "TEXT NOT NULL DEFAULT 'user'")
    ensure_column(connection, "users", "permissions_json", "TEXT NOT NULL DEFAULT '[]'")
    ensure_column(connection, "users", "active", "INTEGER NOT NULL DEFAULT 1")
    ensure_column(connection, "users", "can_create_databases", "INTEGER NOT NULL DEFAULT 0")
    connection.execute("UPDATE users SET role = COALESCE(NULLIF(role, ''), 'user')")
    connection.execute("UPDATE users SET permissions_json = COALESCE(NULLIF(permissions_json, ''), '[]')")
    connection.execute("UPDATE users SET active = COALESCE(active, 1)")
    connection.execute("UPDATE users SET can_create_databases = COALESCE(can_create_databases, 0)")
    connection.execute("UPDATE users SET can_create_databases = 1 WHERE LOWER(COALESCE(role, '')) = 'admin'")
    connection.execute("UPDATE users SET department = 'Maszynownia' WHERE department = 'Planowanie'")
    user_department_placeholders = ",".join("?" for _ in DEPARTMENTS)
    connection.execute(
        f"""
        UPDATE users
        SET department = 'Maszynownia'
        WHERE COALESCE(department, '') = ''
           OR department NOT IN ({user_department_placeholders})
        """,
        tuple(DEPARTMENTS),
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS orders (
          id TEXT PRIMARY KEY,
          order_number TEXT NOT NULL,
          entry_date TEXT NOT NULL,
          material_date TEXT NOT NULL,
          owner TEXT NOT NULL,
          client TEXT NOT NULL,
          color TEXT,
          frames_count INTEGER NOT NULL DEFAULT 0,
          sashes_count INTEGER NOT NULL DEFAULT 0,
          extras TEXT,
          created_at TEXT NOT NULL
        )
        """
    )
    ensure_column(connection, "orders", "order_status", "TEXT NOT NULL DEFAULT 'Dokumentacja'")
    ensure_column(connection, "orders", "archived", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "orders", "archived_at", "TEXT")
    ensure_column(connection, "orders", "completed_at", "TEXT")
    ensure_column(connection, "orders", "manual_start_date", "TEXT")
    ensure_column(connection, "orders", "manual_planned_date", "TEXT")
    connection.execute("UPDATE orders SET order_status = 'Dokumentacja' WHERE order_status = 'Planowanie'")
    connection.execute("UPDATE orders SET archived = 0 WHERE archived IS NULL")
    connection.execute(
        """
        UPDATE orders
        SET archived_at = COALESCE(NULLIF(archived_at, ''), substr(created_at, 1, 10))
        WHERE archived = 1 AND (archived_at IS NULL OR archived_at = '')
        """
    )
    connection.execute(
        """
        UPDATE orders
        SET completed_at = COALESCE(NULLIF(completed_at, ''), archived_at, substr(created_at, 1, 10))
        WHERE order_status = 'Zakonczone' AND (completed_at IS NULL OR completed_at = '')
        """
    )
    connection.execute(
        """
        UPDATE orders
        SET completed_at = NULL
        WHERE order_status <> 'Zakonczone' AND completed_at IS NOT NULL AND completed_at <> ''
        """
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS order_positions (
          id TEXT PRIMARY KEY,
          order_id TEXT NOT NULL,
          position_number TEXT NOT NULL,
          width INTEGER NOT NULL,
          height INTEGER NOT NULL,
          technology TEXT NOT NULL,
          line TEXT NOT NULL,
          machining_time REAL NOT NULL,
          painting_time REAL NOT NULL,
          assembly_time REAL NOT NULL,
          progress_percent REAL NOT NULL DEFAULT 0,
          status TEXT NOT NULL,
          started_at TEXT,
          finished_at TEXT,
          FOREIGN KEY(order_id) REFERENCES orders(id) ON DELETE CASCADE
        )
        """
    )

    ensure_column(connection, "order_positions", "material_wood_date", "TEXT")
    ensure_column(connection, "order_positions", "material_corpus_date", "TEXT")
    ensure_column(connection, "order_positions", "material_glass_date", "TEXT")
    ensure_column(connection, "order_positions", "material_hardware_date", "TEXT")
    ensure_column(connection, "order_positions", "material_accessories_date", "TEXT")
    ensure_column(connection, "order_positions", "material_wood_to_order", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "material_corpus_to_order", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "material_glass_to_order", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "material_hardware_to_order", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "material_accessories_to_order", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "notes", "TEXT")
    ensure_column(connection, "order_positions", "current_department_status", "TEXT NOT NULL DEFAULT 'Dokumentacja'")
    ensure_column(connection, "order_positions", "position_frames_count", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "position_sashes_count", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "shape_rect", "INTEGER NOT NULL DEFAULT 1")
    ensure_column(connection, "order_positions", "shape_skos", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "shape_luk", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "slemie_count", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "slupek_staly_count", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "przymyk_count", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "niski_prog_count", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(connection, "order_positions", "attachment_name", "TEXT")
    ensure_column(connection, "order_positions", "attachment_mime", "TEXT")
    ensure_column(connection, "order_positions", "attachment_data", "TEXT")
    ensure_column(connection, "order_positions", "progress_percent", "REAL NOT NULL DEFAULT 0")
    connection.execute(
        "UPDATE order_positions SET current_department_status = 'Dokumentacja' WHERE current_department_status = 'Planowanie'"
    )
    connection.execute(
        """
        UPDATE order_positions
        SET progress_percent = 100
        WHERE status = 'done' AND (progress_percent IS NULL OR progress_percent < 100)
        """
    )
    connection.execute(
        """
        UPDATE order_positions
        SET progress_percent = 1
        WHERE status = 'in_progress' AND (progress_percent IS NULL OR progress_percent <= 0)
        """
    )
    connection.execute(
        """
        UPDATE order_positions
        SET progress_percent = 0
        WHERE status = 'pending' AND progress_percent IS NULL
        """
    )
    sync_all_order_statuses(connection)

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS feedback_events (
          id TEXT PRIMARY KEY,
          order_id TEXT NOT NULL,
          position_id TEXT NOT NULL,
          action TEXT NOT NULL,
          actor TEXT NOT NULL,
          at TEXT NOT NULL
        )
        """
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS technology_allocations (
          technology_name TEXT PRIMARY KEY,
          data_json TEXT NOT NULL
        )
        """
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS stations (
          id TEXT PRIMARY KEY,
          name TEXT NOT NULL,
          department TEXT NOT NULL,
          active INTEGER NOT NULL DEFAULT 1,
          sort_order INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    ensure_column(connection, "stations", "active", "INTEGER NOT NULL DEFAULT 1")
    ensure_column(connection, "stations", "sort_order", "INTEGER NOT NULL DEFAULT 0")
    connection.execute("UPDATE stations SET active = 1 WHERE active IS NULL")

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS station_settings (
          station_id TEXT PRIMARY KEY,
          shift_count INTEGER NOT NULL,
          people_count INTEGER NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS material_rules (
          id INTEGER PRIMARY KEY CHECK (id = 1),
          data_json TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS skill_workers (
          id TEXT PRIMARY KEY,
          name TEXT NOT NULL,
          department TEXT NOT NULL,
          primary_station_id TEXT,
          active INTEGER NOT NULL DEFAULT 1
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS skill_worker_skills (
          worker_id TEXT NOT NULL,
          station_id TEXT NOT NULL,
          level INTEGER NOT NULL DEFAULT 0,
          PRIMARY KEY (worker_id, station_id),
          FOREIGN KEY(worker_id) REFERENCES skill_workers(id) ON DELETE CASCADE
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS skill_worker_availability (
          worker_id TEXT NOT NULL,
          day TEXT NOT NULL,
          shift INTEGER NOT NULL,
          minutes INTEGER NOT NULL DEFAULT 0,
          PRIMARY KEY (worker_id, day, shift),
          FOREIGN KEY(worker_id) REFERENCES skill_workers(id) ON DELETE CASCADE
        )
        """
    )
    ensure_column(connection, "skill_workers", "active", "INTEGER NOT NULL DEFAULT 1")
    ensure_column(connection, "skill_workers", "primary_station_id", "TEXT")
    connection.execute("UPDATE skill_workers SET active = COALESCE(active, 1)")
    connection.execute(
        """
        UPDATE skill_workers
        SET primary_station_id = NULL
        WHERE COALESCE(primary_station_id, '') <> ''
          AND primary_station_id NOT IN (SELECT id FROM stations)
        """
    )

    settings_row = connection.execute("SELECT id FROM settings WHERE id = 1").fetchone()
    if not settings_row:
        connection.execute(
            """
            INSERT INTO settings (
              id,
              minutes_per_shift,
              working_days_json,
              calendar_overrides_json,
              weekday_shifts_json,
              station_overtime_json
            )
            VALUES (1, 480, '[1,2,3,4,5]', '{}', '{"0":0,"1":2,"2":2,"3":2,"4":2,"5":2,"6":0}', '{}')
            """
        )

    user_count = connection.execute("SELECT COUNT(*) AS cnt FROM users").fetchone()["cnt"]
    if user_count == 0:
        connection.executemany(
            "INSERT INTO users (id, name, department, role, permissions_json, active, can_create_databases) VALUES (?, ?, ?, ?, ?, ?, ?)",
            [
                (str(uuid.uuid4()), "Jan Brygadzista", "Maszynownia", "user", json.dumps(["feedback"], ensure_ascii=True), 1, 0),
                (
                    str(uuid.uuid4()),
                    "Anna Planistka",
                    "Dokumentacja",
                    "user",
                    json.dumps(["orders", "gantt", "reports", "execution"], ensure_ascii=True),
                    1,
                    0,
                ),
            ],
        )

    admin_row = connection.execute(
        """
        SELECT id, name, password_hash, role, permissions_json, active, can_create_databases
        FROM users
        WHERE LOWER(COALESCE(login, '')) = 'admin'
        LIMIT 1
        """
    ).fetchone()
    if not admin_row:
        connection.execute(
            """
            INSERT INTO users (id, name, department, login, password_hash, role, permissions_json, active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(uuid.uuid4()),
                "Administrator",
                "Dokumentacja",
                "admin",
                hash_password("admin"),
                "admin",
                json.dumps(USER_VISIBLE_SECTIONS, ensure_ascii=True),
                1,
            ),
        )
    else:
        admin_name = str(admin_row["name"] or "").strip() or "Administrator"
        admin_hash = str(admin_row["password_hash"] or "").strip() or hash_password("admin")
        admin_role = "admin"
        try:
            admin_sections = normalize_visible_sections(json.loads(admin_row["permissions_json"] or "[]"))
        except json.JSONDecodeError:
            admin_sections = []
        if len(admin_sections) == 0:
            admin_sections = USER_VISIBLE_SECTIONS[:]
        connection.execute(
            """
            UPDATE users
            SET name = ?, role = ?, password_hash = ?, permissions_json = ?, active = 1, can_create_databases = 1
            WHERE id = ?
            """,
            (
                admin_name,
                admin_role,
                admin_hash,
                json.dumps(admin_sections, ensure_ascii=True),
                admin_row["id"],
            ),
        )
    connection.execute("UPDATE users SET can_create_databases = 1 WHERE LOWER(COALESCE(login, '')) = 'admin'")

    tech_count = connection.execute("SELECT COUNT(*) AS cnt FROM technology_allocations").fetchone()["cnt"]
    if tech_count == 0:
        for name, payload in default_technology_allocations().items():
            connection.execute(
                "INSERT INTO technology_allocations (technology_name, data_json) VALUES (?, ?)",
                (name, json.dumps(payload, ensure_ascii=True)),
            )

    material_rules_row = connection.execute("SELECT id FROM material_rules WHERE id = 1").fetchone()
    if not material_rules_row:
        connection.execute(
            "INSERT INTO material_rules (id, data_json) VALUES (1, ?)",
            (json.dumps(DEFAULT_MATERIAL_RULES, ensure_ascii=True),),
        )

    station_count = connection.execute("SELECT COUNT(*) AS cnt FROM stations").fetchone()["cnt"]
    if station_count == 0:
        connection.executemany(
            "INSERT INTO stations (id, name, department, active, sort_order) VALUES (?, ?, ?, ?, ?)",
            [(item["id"], item["name"], item["department"], 1, index + 1) for index, item in enumerate(STATIONS)],
        )

    ordering_rows = connection.execute(
        "SELECT id, sort_order, department, name FROM stations ORDER BY sort_order, department, name, id"
    ).fetchall()
    used_sort_orders = set()
    needs_reindex = False
    for row in ordering_rows:
        value = int(row["sort_order"] or 0)
        if value <= 0 or value in used_sort_orders:
            needs_reindex = True
            break
        used_sort_orders.add(value)
    if needs_reindex:
        for index, row in enumerate(ordering_rows):
            connection.execute("UPDATE stations SET sort_order = ? WHERE id = ?", (index + 1, row["id"]))

    station_rows = connection.execute("SELECT id FROM stations ORDER BY id").fetchall()
    for row in station_rows:
        station_id = row["id"]
        exists = connection.execute(
            "SELECT station_id FROM station_settings WHERE station_id = ?",
            (station_id,),
        ).fetchone()
        if not exists:
            connection.execute(
                "INSERT INTO station_settings (station_id, shift_count, people_count) VALUES (?, ?, ?)",
                (station_id, 2, 1),
            )

    sync_technology_allocations_with_stations(connection)
    connection.commit()
    connection.close()


def get_settings(connection):
    row = connection.execute(
        """
        SELECT
          minutes_per_shift,
          working_days_json,
          calendar_overrides_json,
          weekday_shifts_json,
          station_overtime_json
        FROM settings
        WHERE id = 1
        """
    ).fetchone()
    data = dict(row)
    try:
        working_days = normalize_working_days(json.loads(data.get("working_days_json") or "[]"))
    except json.JSONDecodeError:
        working_days = [1, 2, 3, 4, 5]
    try:
        parsed_shifts = json.loads(data.get("weekday_shifts_json") or "{}")
    except json.JSONDecodeError:
        parsed_shifts = {}
    default_working_days = [1, 2, 3, 4, 5]
    if parsed_shifts == DEFAULT_WEEKDAY_SHIFTS and sorted(working_days) != default_working_days:
        weekday_shifts = normalize_weekday_shifts({}, working_days)
    else:
        weekday_shifts = normalize_weekday_shifts(parsed_shifts, working_days)
    days_from_shifts = [day for day in range(7) if int(weekday_shifts.get(str(day), 0)) > 0]
    data["weekday_shifts"] = weekday_shifts
    data["working_days"] = days_from_shifts or working_days
    try:
        parsed = json.loads(data.get("calendar_overrides_json") or "{}")
    except json.JSONDecodeError:
        parsed = {}
    data["calendar_overrides"] = normalize_calendar_overrides(parsed)
    try:
        parsed_overtime = json.loads(data.get("station_overtime_json") or "{}")
    except json.JSONDecodeError:
        parsed_overtime = {}
    data["station_overtime"] = normalize_station_overtime(parsed_overtime)
    return data


def user_payload_from_row(row):
    if not row:
        return None
    data = dict(row)
    try:
        visible_sections = normalize_visible_sections(json.loads(data.get("permissions_json") or "[]"))
    except json.JSONDecodeError:
        visible_sections = []
    return {
        "id": data["id"],
        "name": data["name"],
        "department": data["department"],
        "login": data.get("login") or "",
        "role": data.get("role") or "user",
        "visibleSections": visible_sections,
        "active": bool(data.get("active", 1)),
        "canCreateDatabases": bool(data.get("can_create_databases", 0)),
    }


def get_active_user_by_id(connection, user_id):
    if not user_id:
        return None
    row = connection.execute(
        """
        SELECT id, name, department, login, role, permissions_json, active, can_create_databases
        FROM users
        WHERE id = ? AND active = 1
        LIMIT 1
        """,
        (user_id,),
    ).fetchone()
    return user_payload_from_row(row)


def get_users(connection):
    rows = connection.execute(
        "SELECT id, name, department, login, role, permissions_json, active, can_create_databases FROM users ORDER BY name"
    ).fetchall()
    return [user_payload_from_row(item) for item in rows]


def get_user_auth_row(connection, login):
    login_value = str(login or "").strip()
    if not login_value:
        return None
    return connection.execute(
        """
        SELECT id, name, department, login, password_hash, role, permissions_json, active, can_create_databases
        FROM users
        WHERE LOWER(COALESCE(login, '')) = LOWER(?)
        LIMIT 1
        """,
        (login_value,),
    ).fetchone()


def get_user_auth_row_from_path(db_path, login):
    path = Path(db_path)
    if not path.exists():
        return None
    initialize(path)
    connection = db_connect(path)
    row = get_user_auth_row(connection, login)
    connection.close()
    return row


def upsert_user_row_to_database(target_db_path, source_row):
    if not source_row:
        return False
    target_path = Path(target_db_path)
    if not target_path.exists():
        return False
    source = dict(source_row)
    login_value = str(source.get("login") or "").strip()
    if not login_value:
        return False
    initialize(target_path)
    connection = db_connect(target_path)
    existing_by_login = get_user_auth_row(connection, login_value)
    if existing_by_login:
        connection.execute(
            """
            UPDATE users
            SET name = ?, department = ?, password_hash = ?, role = ?, permissions_json = ?, active = ?, can_create_databases = ?
            WHERE id = ?
            """,
            (
                str(source.get("name") or "").strip() or "Uzytkownik",
                normalize_standard_department(source.get("department")),
                str(source.get("password_hash") or "").strip() or hash_password("admin"),
                str(source.get("role") or "user").strip().lower() if str(source.get("role") or "").strip() else "user",
                str(source.get("permissions_json") or "[]"),
                1 if bool(source.get("active", 1)) else 0,
                1 if bool(source.get("can_create_databases", 0)) else 0,
                existing_by_login["id"],
            ),
        )
    else:
        source_id = str(source.get("id") or uuid.uuid4())
        id_conflict = connection.execute("SELECT id FROM users WHERE id = ? LIMIT 1", (source_id,)).fetchone()
        if id_conflict:
            source_id = str(uuid.uuid4())
        connection.execute(
            """
            INSERT INTO users (id, name, department, login, password_hash, role, permissions_json, active, can_create_databases)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_id,
                str(source.get("name") or "").strip() or "Uzytkownik",
                normalize_standard_department(source.get("department")),
                login_value,
                str(source.get("password_hash") or "").strip() or hash_password("admin"),
                str(source.get("role") or "user").strip().lower() if str(source.get("role") or "").strip() else "user",
                str(source.get("permissions_json") or "[]"),
                1 if bool(source.get("active", 1)) else 0,
                1 if bool(source.get("can_create_databases", 0)) else 0,
            ),
        )
    connection.execute("UPDATE users SET can_create_databases = 1 WHERE LOWER(COALESCE(role, '')) = 'admin'")
    connection.commit()
    connection.close()
    return True


def get_feedback_events(connection):
    rows = connection.execute(
        """
        SELECT id, order_id AS orderId, position_id AS positionId, action, actor AS by, at
        FROM feedback_events
        ORDER BY at DESC
        """
    ).fetchall()
    return [dict(item) for item in rows]


def get_technology_allocations(connection):
    rows = connection.execute(
        "SELECT technology_name, data_json FROM technology_allocations ORDER BY technology_name"
    ).fetchall()
    output = {}
    for row in rows:
        output[row["technology_name"]] = json.loads(row["data_json"])
    return output


def get_stations(connection):
    rows = connection.execute(
        "SELECT id, name, department, active, sort_order FROM stations ORDER BY sort_order, department, name, id"
    ).fetchall()
    return [
        {
            "id": item["id"],
            "name": item["name"],
            "department": item["department"],
            "active": bool(item["active"]),
            "sortOrder": int(item["sort_order"] or 0),
        }
        for item in rows
    ]


def get_station_settings(connection):
    rows = connection.execute(
        "SELECT station_id, shift_count, people_count FROM station_settings ORDER BY station_id"
    ).fetchall()
    output = {}
    for row in rows:
        output[row["station_id"]] = {
            "shiftCount": row["shift_count"],
            "peopleCount": row["people_count"],
        }
    return output


def get_material_rules(connection):
    row = connection.execute("SELECT data_json FROM material_rules WHERE id = 1").fetchone()
    if not row:
        return DEFAULT_MATERIAL_RULES
    try:
        parsed = json.loads(row["data_json"])
    except json.JSONDecodeError:
        parsed = {}
    result = {dept: [] for dept in DEPARTMENTS}
    for dept, materials in parsed.items():
        if dept not in result or not isinstance(materials, list):
            continue
        result[dept] = [item for item in materials if item in MATERIAL_KEYS]
    return result


def get_orders(connection):
    order_rows = connection.execute(
        """
        SELECT id, order_number, entry_date, material_date, owner, client, color, frames_count, sashes_count, extras, created_at, order_status, archived, archived_at, completed_at, manual_start_date, manual_planned_date
        FROM orders
        ORDER BY created_at DESC
        """
    ).fetchall()

    position_rows = connection.execute(
        """
        SELECT
          id, order_id, position_number, width, height, technology, line,
          machining_time, painting_time, assembly_time,
          progress_percent,
          position_frames_count, position_sashes_count,
          shape_rect, shape_skos, shape_luk,
          slemie_count, slupek_staly_count, przymyk_count, niski_prog_count,
          material_wood_date, material_corpus_date, material_glass_date, material_hardware_date, material_accessories_date,
          material_wood_to_order, material_corpus_to_order, material_glass_to_order, material_hardware_to_order, material_accessories_to_order,
          notes, current_department_status, attachment_name,
          status, started_at, finished_at
        FROM order_positions
        ORDER BY order_id, position_number
        """
    ).fetchall()

    by_order = {}
    for row in position_rows:
        item = dict(row)
        mapped = {
            "id": item["id"],
            "positionNumber": item["position_number"],
            "width": item["width"],
            "height": item["height"],
            "technology": item["technology"],
            "line": item["line"],
            "framesCount": int(item["position_frames_count"] or 0),
            "sashesCount": int(item["position_sashes_count"] or 0),
            "shapeRect": bool(item["shape_rect"]),
            "shapeSkos": bool(item["shape_skos"]),
            "shapeLuk": bool(item["shape_luk"]),
            "slemieCount": int(item["slemie_count"] or 0),
            "slupekStalyCount": int(item["slupek_staly_count"] or 0),
            "przymykCount": int(item["przymyk_count"] or 0),
            "niskiProgCount": int(item["niski_prog_count"] or 0),
            "times": {
                "machining": item["machining_time"],
                "painting": item["painting_time"],
                "assembly": item["assembly_time"],
            },
            "materials": {
                "wood": {
                    "date": item["material_wood_date"],
                    "toOrder": bool(item["material_wood_to_order"]),
                },
                "corpus": {
                    "date": item["material_corpus_date"],
                    "toOrder": bool(item["material_corpus_to_order"]),
                },
                "glass": {
                    "date": item["material_glass_date"],
                    "toOrder": bool(item["material_glass_to_order"]),
                },
                "hardware": {
                    "date": item["material_hardware_date"],
                    "toOrder": bool(item["material_hardware_to_order"]),
                },
                "accessories": {
                    "date": item["material_accessories_date"],
                    "toOrder": bool(item["material_accessories_to_order"]),
                },
            },
            "notes": item["notes"] or "",
            "currentDepartmentStatus": normalize_position_department_for_status(
                item.get("status"), item.get("current_department_status")
            ),
            "progressPercent": normalize_position_progress(item.get("status"), item.get("progress_percent")),
            "attachmentName": item["attachment_name"],
            "status": item["status"],
            "startedAt": item["started_at"],
            "finishedAt": item["finished_at"],
        }
        by_order.setdefault(item["order_id"], []).append(mapped)

    output = []
    for row in order_rows:
        item = dict(row)
        positions = by_order.get(item["id"], [])
        frames_count = sum(max(0, int(position.get("framesCount", 0) or 0)) for position in positions)
        sashes_count = sum(max(0, int(position.get("sashesCount", 0) or 0)) for position in positions)
        output.append(
            {
                "id": item["id"],
                "orderNumber": item["order_number"],
                "entryDate": item["entry_date"],
                "materialDate": item["material_date"],
                "owner": item["owner"],
                "client": item["client"],
                "color": item["color"] or "",
                "framesCount": frames_count,
                "sashesCount": sashes_count,
                "extras": item["extras"] or "",
                "orderStatus": normalize_order_status(item["order_status"]),
                "archived": bool(item["archived"]),
                "archivedAt": item["archived_at"],
                "completedAt": item["completed_at"],
                "manualStartDate": item["manual_start_date"],
                "manualPlannedDate": item["manual_planned_date"],
                "positions": positions,
                "plannedProductionDate": None,
                "calculation": None,
                "createdAt": item["created_at"],
            }
        )
    return output


def get_skill_workers(connection):
    worker_rows = connection.execute(
        """
        SELECT id, name, department, primary_station_id, active
        FROM skill_workers
        ORDER BY department, name, id
        """
    ).fetchall()
    skill_rows = connection.execute(
        """
        SELECT worker_id, station_id, level
        FROM skill_worker_skills
        ORDER BY worker_id, station_id
        """
    ).fetchall()
    skills_by_worker = {}
    for row in skill_rows:
        level = clamp_int(row["level"], 0, 3, 0)
        if level <= 0:
            continue
        wid = str(row["worker_id"] or "")
        if not wid:
            continue
        if wid not in skills_by_worker:
            skills_by_worker[wid] = {}
        skills_by_worker[wid][str(row["station_id"] or "")] = level

    output = []
    for row in worker_rows:
        worker_id = str(row["id"] or "")
        output.append(
            {
                "id": worker_id,
                "name": str(row["name"] or "").strip(),
                "department": str(row["department"] or "").strip() or "Maszynownia",
                "primaryStationId": str(row["primary_station_id"] or "").strip(),
                "active": bool(row["active"]),
                "skills": skills_by_worker.get(worker_id, {}),
            }
        )
    return output


def get_skill_worker_availability(connection):
    rows = connection.execute(
        """
        SELECT worker_id, day, shift, minutes
        FROM skill_worker_availability
        ORDER BY day, shift, worker_id
        """
    ).fetchall()
    output = []
    for row in rows:
        worker_id = str(row["worker_id"] or "").strip()
        day = str(row["day"] or "").strip()
        shift = clamp_int(row["shift"], 1, 3, 1)
        if not worker_id or not is_valid_iso_date(day):
            continue
        minutes = clamp_int(row["minutes"], 0, 1440, 0)
        output.append(
            {
                "workerId": worker_id,
                "date": day,
                "shift": shift,
                "minutes": minutes,
            }
        )
    return output


def normalize_excel_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def parse_excel_sheet_rows(sheet):
    if sheet is None:
        return []
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [normalize_excel_header(item) for item in header_row]
    output = []
    for row_index, row in enumerate(rows_iter, start=2):
        if row is None:
            continue
        if all(value is None or str(value).strip() == "" for value in row):
            continue
        item = {"_rowNumber": row_index}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[col_index] if col_index < len(row) else None
        output.append(item)
    return output


def find_excel_sheet(workbook, aliases):
    if not workbook:
        return None
    normalized = {normalize_excel_header(name): workbook[name] for name in workbook.sheetnames}
    for alias in aliases:
        found = normalized.get(normalize_excel_header(alias))
        if found:
            return found
    return None


def excel_text(value):
    if value is None:
        return ""
    return str(value).strip()


def excel_date_to_iso(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def excel_bool(value, fallback=False):
    if value is None:
        return bool(fallback)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in ("1", "true", "tak", "yes", "y", "on", "x"):
        return True
    if text in ("0", "false", "nie", "no", "off", ""):
        return False
    return bool(fallback)


def excel_int(value, fallback=0):
    try:
        if value is None or str(value).strip() == "":
            return int(fallback)
        return int(float(value))
    except (TypeError, ValueError):
        return int(fallback)


def excel_float(value, fallback=0.0):
    try:
        if value is None or str(value).strip() == "":
            return float(fallback)
        return float(value)
    except (TypeError, ValueError):
        return float(fallback)


def derive_timestamps_for_workflow(workflow_status, existing_started=None, existing_finished=None):
    status = normalize_workflow_status(workflow_status, "pending")
    timestamp = now_iso()
    if status == "done":
        started = normalize_optional_text(existing_started) or timestamp
        finished = normalize_optional_text(existing_finished) or timestamp
        return started, finished
    if status == "in_progress":
        started = normalize_optional_text(existing_started) or timestamp
        return started, None
    return None, None


def upsert_from_excel_workbook(connection, workbook):
    orders_sheet = find_excel_sheet(workbook, ["Zamowienia", "Orders"])
    positions_sheet = find_excel_sheet(workbook, ["Pozycje", "Positions"])
    if not orders_sheet and not positions_sheet:
        raise ValueError("Brak arkuszy 'Zamowienia' i 'Pozycje' w pliku.")

    order_rows = parse_excel_sheet_rows(orders_sheet) if orders_sheet else []
    position_rows = parse_excel_sheet_rows(positions_sheet) if positions_sheet else []
    if not order_rows and not position_rows:
        raise ValueError("Plik nie zawiera danych do importu.")

    existing_orders_rows = connection.execute(
        """
        SELECT id, order_number, entry_date, owner, client, color, extras, order_status, manual_start_date, manual_planned_date
        FROM orders
        """
    ).fetchall()
    existing_orders_by_number = {row["order_number"]: dict(row) for row in existing_orders_rows}

    created_orders = 0
    updated_orders = 0
    created_positions = 0
    updated_positions = 0
    skipped_rows = 0
    errors = []
    order_id_by_number = {}
    touched_order_ids = set()

    for row in order_rows:
        row_no = int(row.get("_rowNumber") or 0)
        order_number = excel_text(row.get("ordernumber"))
        if not order_number:
            skipped_rows += 1
            errors.append(f"Zamowienia, wiersz {row_no}: brak numeru zamowienia.")
            continue

        existing = existing_orders_by_number.get(order_number)
        entry_date = excel_date_to_iso(row.get("entrydate"))
        owner = excel_text(row.get("owner"))
        client = excel_text(row.get("client"))
        color = excel_text(row.get("color"))
        extras = excel_text(row.get("extras"))
        order_status_raw = excel_text(row.get("orderstatus"))
        order_status = normalize_order_status(order_status_raw or (existing.get("order_status") if existing else "Dokumentacja"))
        manual_start_date = excel_date_to_iso(row.get("manualstartdate"))
        manual_planned_date = excel_date_to_iso(row.get("manualplanneddate"))

        if existing:
            entry_date = entry_date or existing.get("entry_date")
            owner = owner or excel_text(existing.get("owner"))
            client = client or excel_text(existing.get("client"))
            color = color if color != "" else excel_text(existing.get("color"))
            extras = extras if extras != "" else excel_text(existing.get("extras"))
            manual_start_date = manual_start_date or normalize_optional_date(existing.get("manual_start_date"))
            manual_planned_date = manual_planned_date or normalize_optional_date(existing.get("manual_planned_date"))
        else:
            if not entry_date:
                skipped_rows += 1
                errors.append(f"Zamowienia, wiersz {row_no}: brak lub zla data wejscia.")
                continue
            if not owner or not client:
                skipped_rows += 1
                errors.append(f"Zamowienia, wiersz {row_no}: brak pola opracowuje lub klient.")
                continue

        if not entry_date or not owner or not client:
            skipped_rows += 1
            errors.append(f"Zamowienia, wiersz {row_no}: niekompletne dane po uzupelnieniu.")
            continue

        if existing:
            connection.execute(
                """
                UPDATE orders
                SET entry_date = ?, material_date = ?, owner = ?, client = ?, color = ?, extras = ?, order_status = ?,
                    manual_start_date = ?, manual_planned_date = ?,
                    completed_at = CASE
                      WHEN ? = 'Zakonczone' THEN COALESCE(NULLIF(completed_at, ''), DATE('now', 'localtime'))
                      ELSE NULL
                    END
                WHERE id = ?
                """,
                (
                    entry_date,
                    entry_date,
                    owner,
                    client,
                    color,
                    extras,
                    order_status,
                    manual_start_date,
                    manual_planned_date,
                    order_status,
                    existing["id"],
                ),
            )
            order_id = existing["id"]
            updated_orders += 1
        else:
            order_id = str(uuid.uuid4())
            connection.execute(
                """
                INSERT INTO orders (id, order_number, entry_date, material_date, owner, client, color, frames_count, sashes_count, extras, created_at, order_status, archived, archived_at, completed_at, manual_start_date, manual_planned_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    order_id,
                    order_number,
                    entry_date,
                    entry_date,
                    owner,
                    client,
                    color,
                    0,
                    0,
                    extras,
                    now_iso(),
                    order_status,
                    0,
                    None,
                    datetime.now().date().isoformat() if order_status == "Zakonczone" else None,
                    manual_start_date,
                    manual_planned_date,
                ),
            )
            created_orders += 1

        existing_orders_by_number[order_number] = {
            "id": order_id,
            "order_number": order_number,
            "entry_date": entry_date,
            "owner": owner,
            "client": client,
            "color": color,
            "extras": extras,
            "order_status": order_status,
            "manual_start_date": manual_start_date,
            "manual_planned_date": manual_planned_date,
        }
        order_id_by_number[order_number] = order_id
        touched_order_ids.add(order_id)

    for order_number, row in existing_orders_by_number.items():
        order_id_by_number.setdefault(order_number, row["id"])

    if order_id_by_number:
        placeholders = ",".join("?" for _ in order_id_by_number)
        existing_position_rows = connection.execute(
            f"""
            SELECT id, order_id, position_number, started_at, finished_at
            FROM order_positions
            WHERE order_id IN (
              SELECT id FROM orders WHERE order_number IN ({placeholders})
            )
            """,
            list(order_id_by_number.keys()),
        ).fetchall()
    else:
        existing_position_rows = []
    existing_positions_by_key = {
        (row["order_id"], str(row["position_number"]).strip()): dict(row) for row in existing_position_rows
    }

    for row in position_rows:
        row_no = int(row.get("_rowNumber") or 0)
        order_number = excel_text(row.get("ordernumber"))
        position_number = excel_text(row.get("positionnumber"))
        technology = excel_text(row.get("technology"))
        line = excel_text(row.get("line"))
        width = excel_int(row.get("width"), 0)
        height = excel_int(row.get("height"), 0)
        if not order_number or not position_number:
            skipped_rows += 1
            errors.append(f"Pozycje, wiersz {row_no}: brak numeru zamowienia lub pozycji.")
            continue
        if not technology or not line:
            skipped_rows += 1
            errors.append(f"Pozycje, wiersz {row_no}: brak technologii lub linii.")
            continue
        if width <= 0 or height <= 0:
            skipped_rows += 1
            errors.append(f"Pozycje, wiersz {row_no}: wymiary musza byc > 0.")
            continue
        order_id = order_id_by_number.get(order_number)
        if not order_id:
            skipped_rows += 1
            errors.append(f"Pozycje, wiersz {row_no}: nie znaleziono zamowienia '{order_number}'.")
            continue

        machining = excel_float(row.get("machiningtime"), 0.0)
        painting = excel_float(row.get("paintingtime"), 0.0)
        assembly = excel_float(row.get("assemblytime"), 0.0)
        workflow_status = normalize_workflow_status(excel_text(row.get("workflowstatus")) or "pending", "pending")
        department_status = normalize_department_status(excel_text(row.get("currentdepartmentstatus")) or "Dokumentacja")
        progress_percent = normalize_position_progress(workflow_status, parse_progress_percent(row.get("progresspercent")))

        materials = {
            "wood": (excel_date_to_iso(row.get("materialwooddate")), excel_bool(row.get("materialwoodtoorder"), False)),
            "corpus": (excel_date_to_iso(row.get("materialcorpusdate")), excel_bool(row.get("materialcorpustoorder"), False)),
            "glass": (excel_date_to_iso(row.get("materialglassdate")), excel_bool(row.get("materialglasstoorder"), False)),
            "hardware": (
                excel_date_to_iso(row.get("materialhardwaredate")),
                excel_bool(row.get("materialhardwaretoorder"), False),
            ),
            "accessories": (
                excel_date_to_iso(row.get("materialaccessoriesdate")),
                excel_bool(row.get("materialaccessoriestoorder"), False),
            ),
        }

        key = (order_id, position_number)
        existing_position = existing_positions_by_key.get(key)
        started_at, finished_at = derive_timestamps_for_workflow(
            workflow_status,
            existing_started=existing_position.get("started_at") if existing_position else None,
            existing_finished=existing_position.get("finished_at") if existing_position else None,
        )

        if existing_position:
            connection.execute(
                """
                UPDATE order_positions
                SET
                  width = ?, height = ?, technology = ?, line = ?,
                  position_frames_count = ?, position_sashes_count = ?,
                  shape_rect = ?, shape_skos = ?, shape_luk = ?,
                  slemie_count = ?, slupek_staly_count = ?, przymyk_count = ?, niski_prog_count = ?,
                  machining_time = ?, painting_time = ?, assembly_time = ?,
                  material_wood_date = ?, material_corpus_date = ?, material_glass_date = ?, material_hardware_date = ?, material_accessories_date = ?,
                  material_wood_to_order = ?, material_corpus_to_order = ?, material_glass_to_order = ?, material_hardware_to_order = ?, material_accessories_to_order = ?,
                  notes = ?, current_department_status = ?, status = ?, progress_percent = ?, started_at = ?, finished_at = ?
                WHERE id = ?
                """,
                (
                    width,
                    height,
                    technology,
                    line,
                    excel_int(row.get("framescount"), 0),
                    excel_int(row.get("sashescount"), 0),
                    int(excel_bool(row.get("shaperect"), True)),
                    int(excel_bool(row.get("shapeskos"), False)),
                    int(excel_bool(row.get("shapeluk"), False)),
                    excel_int(row.get("slemiecount"), 0),
                    excel_int(row.get("slupekstalycount"), 0),
                    excel_int(row.get("przymykcount"), 0),
                    excel_int(row.get("niskiprogcount"), 0),
                    machining,
                    painting,
                    assembly,
                    materials["wood"][0],
                    materials["corpus"][0],
                    materials["glass"][0],
                    materials["hardware"][0],
                    materials["accessories"][0],
                    int(materials["wood"][1]),
                    int(materials["corpus"][1]),
                    int(materials["glass"][1]),
                    int(materials["hardware"][1]),
                    int(materials["accessories"][1]),
                    excel_text(row.get("notes")),
                    department_status,
                    workflow_status,
                    progress_percent,
                    started_at,
                    finished_at,
                    existing_position["id"],
                ),
            )
            updated_positions += 1
        else:
            position_id = str(uuid.uuid4())
            connection.execute(
                """
                INSERT INTO order_positions (
                  id, order_id, position_number, width, height, technology, line,
                  position_frames_count, position_sashes_count,
                  shape_rect, shape_skos, shape_luk,
                  slemie_count, slupek_staly_count, przymyk_count, niski_prog_count,
                  machining_time, painting_time, assembly_time,
                  material_wood_date, material_corpus_date, material_glass_date, material_hardware_date, material_accessories_date,
                  material_wood_to_order, material_corpus_to_order, material_glass_to_order, material_hardware_to_order, material_accessories_to_order,
                  notes, current_department_status, attachment_name, attachment_mime, attachment_data,
                  progress_percent, status, started_at, finished_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    position_id,
                    order_id,
                    position_number,
                    width,
                    height,
                    technology,
                    line,
                    excel_int(row.get("framescount"), 0),
                    excel_int(row.get("sashescount"), 0),
                    int(excel_bool(row.get("shaperect"), True)),
                    int(excel_bool(row.get("shapeskos"), False)),
                    int(excel_bool(row.get("shapeluk"), False)),
                    excel_int(row.get("slemiecount"), 0),
                    excel_int(row.get("slupekstalycount"), 0),
                    excel_int(row.get("przymykcount"), 0),
                    excel_int(row.get("niskiprogcount"), 0),
                    machining,
                    painting,
                    assembly,
                    materials["wood"][0],
                    materials["corpus"][0],
                    materials["glass"][0],
                    materials["hardware"][0],
                    materials["accessories"][0],
                    int(materials["wood"][1]),
                    int(materials["corpus"][1]),
                    int(materials["glass"][1]),
                    int(materials["hardware"][1]),
                    int(materials["accessories"][1]),
                    excel_text(row.get("notes")),
                    department_status,
                    None,
                    None,
                    None,
                    progress_percent,
                    workflow_status,
                    started_at,
                    finished_at,
                ),
            )
            created_positions += 1

        touched_order_ids.add(order_id)

    if touched_order_ids:
        sync_order_statuses_for_orders(connection, sorted(touched_order_ids))

    return {
        "createdOrders": created_orders,
        "updatedOrders": updated_orders,
        "createdPositions": created_positions,
        "updatedPositions": updated_positions,
        "skippedRows": skipped_rows,
        "errorCount": len(errors),
        "errors": errors[:150],
    }


PUBLIC_API_PATHS = {
    "/api/auth/login",
    "/api/auth/logout",
    "/api/auth/session",
    "/api/public/databases",
}


def current_authenticated_user():
    cached = getattr(g, "current_user", None)
    if cached:
        return cached
    user_id = session.get("user_id")
    if not user_id:
        return None
    connection = db()
    user = get_active_user_by_id(connection, user_id)
    connection.close()
    if not user:
        session.pop("user_id", None)
        return None
    g.current_user = user
    return user


@app.before_request
def require_api_authentication():
    if request.method == "OPTIONS":
        return None
    if not request.path.startswith("/api/"):
        return None
    if request.path in PUBLIC_API_PATHS:
        return None
    user = current_authenticated_user()
    if not user:
        return jsonify({"error": "Sesja wygasla. Zaloguj sie ponownie."}), 401
    if not can_user_access_database(user, get_request_database_key()):
        session.pop("user_id", None)
        g.current_user = None
        return jsonify({"error": "Brak dostepu do aktualnej bazy. Zaloguj sie ponownie."}), 403
    return None


@app.get("/")
def index():
    return app.send_static_file("index.html")


@app.get("/api/bootstrap")
def api_bootstrap():
    user = current_authenticated_user()
    connection = db()
    if sync_all_order_statuses(connection):
        connection.commit()
    selected_key = get_request_database_key()
    payload = {
        "orders": get_orders(connection),
        "feedbackEvents": get_feedback_events(connection),
        "users": get_users(connection),
        "settings": get_settings(connection),
        "stations": get_stations(connection),
        "stationSettings": get_station_settings(connection),
        "skillWorkers": get_skill_workers(connection),
        "skillAvailability": get_skill_worker_availability(connection),
        "materialRules": get_material_rules(connection),
        "technologies": get_technology_allocations(connection),
        "databases": list_database_catalog_for_user(user, selected_key),
        "activeDatabase": selected_key,
        "databaseAccessMap": load_database_access_map() if str((user or {}).get("role", "")).strip().lower() == "admin" else {},
    }
    connection.close()
    return jsonify(payload)


@app.post("/api/orders")
def api_create_order():
    data = request.get_json(force=True)
    order_number = str(data.get("orderNumber", "")).strip()
    entry_date = str(data.get("entryDate", "")).strip()
    order_status = normalize_order_status(data.get("orderStatus", "Dokumentacja"))
    owner = str(data.get("owner", "")).strip()
    client = str(data.get("client", "")).strip()
    manual_planned_date = normalize_optional_date(data.get("manualPlannedDate"))
    if not order_number or not entry_date or not owner or not client:
        return jsonify({"error": "UzupeĹ‚nij pola wymagane: numer, data wejĹ›cia, opracowuje, klient."}), 400

    if manual_planned_date and not is_valid_iso_date(manual_planned_date):
        return jsonify({"error": "Nieprawidlowa planowana data produkcji. Oczekiwano formatu RRRR-MM-DD."}), 400

    order_id = str(uuid.uuid4())
    connection = db()

    connection.execute(
        """
        INSERT INTO orders (id, order_number, entry_date, material_date, owner, client, color, frames_count, sashes_count, extras, created_at, order_status, archived, archived_at, completed_at, manual_start_date, manual_planned_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            order_id,
            order_number,
            entry_date,
            entry_date,
            owner,
            client,
            str(data.get("color", "")).strip(),
            int(data.get("framesCount", 0)),
            int(data.get("sashesCount", 0)),
            str(data.get("extras", "")).strip(),
            now_iso(),
            order_status,
            0,
            None,
            (datetime.now().date().isoformat() if order_status == "Zakonczone" else None),
            normalize_optional_date(data.get("manualStartDate")),
            manual_planned_date,
        ),
    )
    connection.commit()
    connection.close()
    return jsonify({"id": order_id}), 201


@app.post("/api/orders/<order_id>/positions")
def api_create_position(order_id):
    data = request.get_json(force=True)
    position_number = str(data.get("positionNumber", "")).strip()
    technology = str(data.get("technology", "")).strip()
    line = str(data.get("line", "")).strip()
    if not position_number or not technology or not line:
        return jsonify({"error": "UzupeĹ‚nij pola pozycji: numer, technologia, linia."}), 400

    position_id = str(uuid.uuid4())
    materials = data.get("materials", {}) or {}
    attachment = data.get("attachment", {}) if isinstance(data.get("attachment"), dict) else {}
    connection = db()
    order_exists = connection.execute("SELECT id FROM orders WHERE id = ?", (order_id,)).fetchone()
    if not order_exists:
        connection.close()
        return jsonify({"error": "Nie znaleziono zamĂłwienia."}), 404

    connection.execute(
        """
        INSERT INTO order_positions (
          id, order_id, position_number, width, height, technology, line,
          position_frames_count, position_sashes_count,
          shape_rect, shape_skos, shape_luk,
          slemie_count, slupek_staly_count, przymyk_count, niski_prog_count,
          machining_time, painting_time, assembly_time,
          material_wood_date, material_corpus_date, material_glass_date, material_hardware_date, material_accessories_date,
          material_wood_to_order, material_corpus_to_order, material_glass_to_order, material_hardware_to_order, material_accessories_to_order,
          notes, current_department_status, attachment_name, attachment_mime, attachment_data,
          progress_percent, status, started_at, finished_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            position_id,
            order_id,
            position_number,
            int(data.get("width", 0)),
            int(data.get("height", 0)),
            technology,
            line,
            int(data.get("framesCount", 0)),
            int(data.get("sashesCount", 0)),
            int(bool(data.get("shapeRect", True))),
            int(bool(data.get("shapeSkos", False))),
            int(bool(data.get("shapeLuk", False))),
            int(data.get("slemieCount", 0)),
            int(data.get("slupekStalyCount", 0)),
            int(data.get("przymykCount", 0)),
            int(data.get("niskiProgCount", 0)),
            float(data.get("times", {}).get("machining", 0)),
            float(data.get("times", {}).get("painting", 0)),
            float(data.get("times", {}).get("assembly", 0)),
            normalize_optional_date(materials.get("wood", {}).get("date")),
            normalize_optional_date(materials.get("corpus", {}).get("date")),
            normalize_optional_date(materials.get("glass", {}).get("date")),
            normalize_optional_date(materials.get("hardware", {}).get("date")),
            normalize_optional_date(materials.get("accessories", {}).get("date")),
            int(bool(materials.get("wood", {}).get("toOrder"))),
            int(bool(materials.get("corpus", {}).get("toOrder"))),
            int(bool(materials.get("glass", {}).get("toOrder"))),
            int(bool(materials.get("hardware", {}).get("toOrder"))),
            int(bool(materials.get("accessories", {}).get("toOrder"))),
            str(data.get("notes", "")).strip(),
            normalize_department_status(data.get("currentDepartmentStatus", "Dokumentacja")),
            normalize_optional_text(attachment.get("name")),
            normalize_optional_text(attachment.get("mimeType")),
            normalize_optional_text(attachment.get("dataBase64")),
            0.0,
            "pending",
            None,
            None,
        ),
    )
    sync_order_status_from_positions(connection, order_id)
    connection.commit()
    connection.close()
    return jsonify({"id": position_id}), 201


@app.post("/api/orders/import-excel")
def api_import_orders_excel():
    if load_workbook is None:
        return jsonify({"error": "Brak biblioteki openpyxl. Zainstaluj zaleznosci aplikacji."}), 500

    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"error": "Wybierz plik Excel do importu."}), 400
    if not str(upload.filename).lower().endswith(".xlsx"):
        return jsonify({"error": "Obslugiwany jest tylko format .xlsx."}), 400

    workbook = None
    connection = None
    try:
        workbook = load_workbook(upload.stream, data_only=True)
        connection = db()
        summary = upsert_from_excel_workbook(connection, workbook)
        connection.commit()
        return jsonify({"ok": True, "summary": summary})
    except ValueError as exc:
        if connection:
            connection.rollback()
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover - defensive runtime guard
        if connection:
            connection.rollback()
        return jsonify({"error": f"Import nie powiodl sie: {exc}"}), 500
    finally:
        if connection:
            connection.close()
        if workbook:
            workbook.close()


@app.put("/api/orders/<order_id>")
def api_update_order(order_id):
    data = request.get_json(force=True)
    order_number = str(data.get("orderNumber", "")).strip()
    entry_date = str(data.get("entryDate", "")).strip()
    order_status = normalize_order_status(data.get("orderStatus", "Dokumentacja"))
    owner = str(data.get("owner", "")).strip()
    client = str(data.get("client", "")).strip()
    manual_planned_date = normalize_optional_date(data.get("manualPlannedDate"))
    if not order_number or not entry_date or not owner or not client:
        return jsonify({"error": "UzupeÄąâ€šnij pola wymagane: numer, data wejÄąâ€şcia, opracowuje, klient."}), 400

    if manual_planned_date and not is_valid_iso_date(manual_planned_date):
        return jsonify({"error": "Nieprawidlowa planowana data produkcji. Oczekiwano formatu RRRR-MM-DD."}), 400

    connection = db()
    row = connection.execute("SELECT id FROM orders WHERE id = ?", (order_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono zamowienia."}), 404

    connection.execute(
        """
        UPDATE orders
        SET order_number = ?, entry_date = ?, material_date = ?, owner = ?, client = ?, color = ?,
            frames_count = ?, sashes_count = ?, extras = ?, order_status = ?,
            completed_at = CASE
              WHEN ? = 'Zakonczone' THEN COALESCE(NULLIF(completed_at, ''), DATE('now', 'localtime'))
              ELSE NULL
            END,
            manual_planned_date = ?
        WHERE id = ?
        """,
        (
            order_number,
            entry_date,
            entry_date,
            owner,
            client,
            str(data.get("color", "")).strip(),
            int(data.get("framesCount", 0)),
            int(data.get("sashesCount", 0)),
            str(data.get("extras", "")).strip(),
            order_status,
            order_status,
            manual_planned_date,
            order_id,
        ),
    )
    sync_order_status_from_positions(connection, order_id)
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/orders/<order_id>/status")
def api_update_order_status(order_id):
    data = request.get_json(force=True)
    order_status = normalize_order_status(data.get("orderStatus", "Dokumentacja"))
    connection = db()
    connection.execute(
        """
        UPDATE orders
        SET order_status = ?,
            completed_at = CASE
              WHEN ? = 'Zakonczone' THEN COALESCE(NULLIF(completed_at, ''), DATE('now', 'localtime'))
              ELSE NULL
            END
        WHERE id = ?
        """,
        (order_status, order_status, order_id),
    )
    sync_order_status_from_positions(connection, order_id)
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/orders/<order_id>/manual-start")
def api_update_order_manual_start(order_id):
    data = request.get_json(force=True)
    manual_start_date = normalize_optional_date(data.get("manualStartDate"))
    if manual_start_date and not is_valid_iso_date(manual_start_date):
        return jsonify({"error": "Nieprawidlowa data recznego startu. Oczekiwano formatu RRRR-MM-DD."}), 400

    connection = db()
    row = connection.execute("SELECT id FROM orders WHERE id = ?", (order_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono zamowienia."}), 404

    connection.execute(
        "UPDATE orders SET manual_start_date = ? WHERE id = ?",
        (manual_start_date, order_id),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.post("/api/orders/archive-completed")
def api_archive_completed_orders():
    data = request.get_json(silent=True) or {}
    raw_ids = data.get("orderIds") if isinstance(data, dict) else None
    selected_ids = []
    if isinstance(raw_ids, list):
        for item in raw_ids:
            value = str(item or "").strip()
            if value and value not in selected_ids:
                selected_ids.append(value)

    connection = db()
    if selected_ids:
        placeholders = ",".join(["?"] * len(selected_ids))
        params = [*selected_ids, "Zakonczone"]
        cursor = connection.execute(
            f"""
            UPDATE orders
            SET archived = 1,
                archived_at = COALESCE(NULLIF(archived_at, ''), DATE('now', 'localtime')),
                completed_at = COALESCE(NULLIF(completed_at, ''), DATE('now', 'localtime'))
            WHERE id IN ({placeholders}) AND order_status = ? AND archived = 0
            """,
            params,
        )
    else:
        cursor = connection.execute(
            """
            UPDATE orders
            SET archived = 1,
                archived_at = COALESCE(NULLIF(archived_at, ''), DATE('now', 'localtime')),
                completed_at = COALESCE(NULLIF(completed_at, ''), DATE('now', 'localtime'))
            WHERE order_status = ? AND archived = 0
            """,
            ("Zakonczone",),
        )
    archived_count = cursor.rowcount or 0
    connection.commit()
    connection.close()
    return jsonify({"ok": True, "archivedCount": archived_count})


@app.put("/api/positions/<position_id>")
def api_update_position(position_id):
    data = request.get_json(force=True)
    connection = db()
    row = connection.execute("SELECT * FROM order_positions WHERE id = ?", (position_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pozycji."}), 404

    item = dict(row)
    update_full = any(
        key in data
        for key in [
            "positionNumber",
            "width",
            "height",
            "technology",
            "line",
            "times",
            "materials",
            "framesCount",
            "sashesCount",
            "shapeRect",
            "shapeSkos",
            "shapeLuk",
            "slemieCount",
            "slupekStalyCount",
            "przymykCount",
            "niskiProgCount",
            "attachment",
            "clearAttachment",
        ]
    )

    if not update_full:
        notes = str(data.get("notes", item.get("notes", ""))).strip()
        if "currentDepartmentStatus" in data:
            current_department_status = normalize_department_status(data.get("currentDepartmentStatus", "Dokumentacja"))
            connection.execute(
                "UPDATE order_positions SET notes = ?, current_department_status = ? WHERE id = ?",
                (notes, current_department_status, position_id),
            )
        else:
            connection.execute(
                "UPDATE order_positions SET notes = ? WHERE id = ?",
                (notes, position_id),
            )
        sync_order_status_from_positions(connection, item["order_id"])
        connection.commit()
        connection.close()
        return jsonify({"ok": True})

    materials = data.get("materials", {}) or {}
    attachment = data.get("attachment") if isinstance(data.get("attachment"), dict) else None
    clear_attachment = to_bool(data.get("clearAttachment"))
    current_department_status = normalize_department_status(
        data.get("currentDepartmentStatus", item.get("current_department_status", "Dokumentacja"))
    )

    attachment_name = item.get("attachment_name")
    attachment_mime = item.get("attachment_mime")
    attachment_data = item.get("attachment_data")
    if attachment is not None:
        attachment_name = normalize_optional_text(attachment.get("name"))
        attachment_mime = normalize_optional_text(attachment.get("mimeType"))
        attachment_data = normalize_optional_text(attachment.get("dataBase64"))
    elif clear_attachment:
        attachment_name = None
        attachment_mime = None
        attachment_data = None

    connection.execute(
        """
        UPDATE order_positions
        SET
          position_number = ?, width = ?, height = ?, technology = ?, line = ?,
          position_frames_count = ?, position_sashes_count = ?,
          shape_rect = ?, shape_skos = ?, shape_luk = ?,
          slemie_count = ?, slupek_staly_count = ?, przymyk_count = ?, niski_prog_count = ?,
          machining_time = ?, painting_time = ?, assembly_time = ?,
          material_wood_date = ?, material_corpus_date = ?, material_glass_date = ?, material_hardware_date = ?, material_accessories_date = ?,
          material_wood_to_order = ?, material_corpus_to_order = ?, material_glass_to_order = ?, material_hardware_to_order = ?, material_accessories_to_order = ?,
          notes = ?, current_department_status = ?, attachment_name = ?, attachment_mime = ?, attachment_data = ?
        WHERE id = ?
        """,
        (
            str(data.get("positionNumber", item.get("position_number", ""))).strip() or item.get("position_number", ""),
            int(data.get("width", item.get("width", 0))),
            int(data.get("height", item.get("height", 0))),
            str(data.get("technology", item.get("technology", ""))).strip() or item.get("technology", ""),
            str(data.get("line", item.get("line", ""))).strip() or item.get("line", ""),
            int(data.get("framesCount", item.get("position_frames_count", 0))),
            int(data.get("sashesCount", item.get("position_sashes_count", 0))),
            int(bool(data.get("shapeRect", bool(item.get("shape_rect", 1))))),
            int(bool(data.get("shapeSkos", bool(item.get("shape_skos", 0))))),
            int(bool(data.get("shapeLuk", bool(item.get("shape_luk", 0))))),
            int(data.get("slemieCount", item.get("slemie_count", 0))),
            int(data.get("slupekStalyCount", item.get("slupek_staly_count", 0))),
            int(data.get("przymykCount", item.get("przymyk_count", 0))),
            int(data.get("niskiProgCount", item.get("niski_prog_count", 0))),
            float((data.get("times") or {}).get("machining", item.get("machining_time", 0))),
            float((data.get("times") or {}).get("painting", item.get("painting_time", 0))),
            float((data.get("times") or {}).get("assembly", item.get("assembly_time", 0))),
            normalize_optional_date((materials.get("wood") or {}).get("date") if "wood" in materials else item.get("material_wood_date")),
            normalize_optional_date((materials.get("corpus") or {}).get("date") if "corpus" in materials else item.get("material_corpus_date")),
            normalize_optional_date((materials.get("glass") or {}).get("date") if "glass" in materials else item.get("material_glass_date")),
            normalize_optional_date((materials.get("hardware") or {}).get("date") if "hardware" in materials else item.get("material_hardware_date")),
            normalize_optional_date(
                (materials.get("accessories") or {}).get("date") if "accessories" in materials else item.get("material_accessories_date")
            ),
            int(
                bool(
                    (materials.get("wood") or {}).get("toOrder")
                    if "wood" in materials
                    else bool(item.get("material_wood_to_order", 0))
                )
            ),
            int(
                bool(
                    (materials.get("corpus") or {}).get("toOrder")
                    if "corpus" in materials
                    else bool(item.get("material_corpus_to_order", 0))
                )
            ),
            int(
                bool(
                    (materials.get("glass") or {}).get("toOrder")
                    if "glass" in materials
                    else bool(item.get("material_glass_to_order", 0))
                )
            ),
            int(
                bool(
                    (materials.get("hardware") or {}).get("toOrder")
                    if "hardware" in materials
                    else bool(item.get("material_hardware_to_order", 0))
                )
            ),
            int(
                bool(
                    (materials.get("accessories") or {}).get("toOrder")
                    if "accessories" in materials
                    else bool(item.get("material_accessories_to_order", 0))
                )
            ),
            str(data.get("notes", item.get("notes", ""))).strip(),
            current_department_status,
            attachment_name,
            attachment_mime,
            attachment_data,
            position_id,
        ),
    )
    sync_order_status_from_positions(connection, item["order_id"])
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/positions/bulk-department")
def api_bulk_position_department():
    data = request.get_json(force=True)
    position_ids = data.get("positionIds", [])
    current_department_status = normalize_department_status(data.get("currentDepartmentStatus", "Dokumentacja"))
    if not isinstance(position_ids, list) or not position_ids:
        return jsonify({"error": "Brak pozycji do aktualizacji."}), 400

    connection = db()
    placeholders = ",".join("?" for _ in position_ids)
    order_rows = connection.execute(
        f"SELECT DISTINCT order_id FROM order_positions WHERE id IN ({placeholders})",
        position_ids,
    ).fetchall()
    order_ids = [row["order_id"] for row in order_rows]
    connection.execute(
        f"UPDATE order_positions SET current_department_status = ? WHERE id IN ({placeholders})",
        [current_department_status, *position_ids],
    )
    sync_order_statuses_for_orders(connection, order_ids)
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/positions/bulk-feedback")
def api_bulk_feedback_update():
    data = request.get_json(force=True)
    position_ids = data.get("positionIds", [])
    if not isinstance(position_ids, list) or not position_ids:
        return jsonify({"error": "Brak pozycji do aktualizacji."}), 400

    has_workflow = "workflowStatus" in data and data.get("workflowStatus") is not None
    has_department = "currentDepartmentStatus" in data and data.get("currentDepartmentStatus") is not None
    has_progress = "progressPercent" in data
    if not has_workflow and not has_department and not has_progress:
        return jsonify({"error": "Brak danych do aktualizacji (status / dzial / procent)."}), 400

    workflow_status = normalize_workflow_status(data.get("workflowStatus"), "pending") if has_workflow else None
    department_status = normalize_department_status(data.get("currentDepartmentStatus")) if has_department else None
    progress_percent = parse_progress_percent(data.get("progressPercent")) if has_progress else None
    actor = str(data.get("actor", "Brygadzista")).strip() or "Brygadzista"

    connection = db()
    placeholders = ",".join("?" for _ in position_ids)
    rows = connection.execute(
        f"""
        SELECT id, order_id, status, current_department_status, progress_percent, started_at, finished_at
        FROM order_positions
        WHERE id IN ({placeholders})
        """,
        position_ids,
    ).fetchall()
    if not rows:
        connection.close()
        return jsonify({"error": "Nie znaleziono pozycji do aktualizacji."}), 404

    changed_count = 0
    order_ids = []
    for row in rows:
        if apply_feedback_update_to_position(
            connection,
            row,
            workflow_status=workflow_status,
            department_status=department_status,
            progress_percent=progress_percent,
            actor=actor,
        ):
            changed_count += 1
        order_ids.append(row["order_id"])

    sync_order_statuses_for_orders(connection, order_ids)
    connection.commit()
    connection.close()
    return jsonify({"ok": True, "updatedCount": changed_count})


@app.put("/api/positions/<position_id>/feedback")
def api_update_position_feedback(position_id):
    data = request.get_json(force=True)
    has_workflow = "workflowStatus" in data and data.get("workflowStatus") is not None
    has_department = "currentDepartmentStatus" in data and data.get("currentDepartmentStatus") is not None
    has_progress = "progressPercent" in data
    if not has_workflow and not has_department and not has_progress:
        return jsonify({"error": "Brak danych do aktualizacji (status / dzial / procent)."}), 400

    workflow_status = normalize_workflow_status(data.get("workflowStatus"), "pending") if has_workflow else None
    department_status = normalize_department_status(data.get("currentDepartmentStatus")) if has_department else None
    progress_percent = parse_progress_percent(data.get("progressPercent")) if has_progress else None
    actor = str(data.get("actor", "Brygadzista")).strip() or "Brygadzista"

    connection = db()
    row = connection.execute(
        """
        SELECT id, order_id, status, current_department_status, progress_percent, started_at, finished_at
        FROM order_positions
        WHERE id = ?
        """,
        (position_id,),
    ).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pozycji."}), 404

    apply_feedback_update_to_position(
        connection,
        row,
        workflow_status=workflow_status,
        department_status=department_status,
        progress_percent=progress_percent,
        actor=actor,
    )
    sync_order_status_from_positions(connection, row["order_id"])
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.get("/api/positions/<position_id>/attachment")
def api_get_position_attachment(position_id):
    connection = db()
    row = connection.execute(
        "SELECT attachment_name, attachment_mime, attachment_data FROM order_positions WHERE id = ?",
        (position_id,),
    ).fetchone()
    connection.close()
    if not row or not row["attachment_data"]:
        return jsonify({"error": "Brak zalacznika."}), 404

    try:
        content = base64.b64decode(row["attachment_data"])
    except (ValueError, TypeError):
        return jsonify({"error": "Nie mozna odczytac zalacznika."}), 500

    filename = row["attachment_name"] or "zalacznik.bin"
    mime = row["attachment_mime"] or "application/octet-stream"
    response = Response(content, mimetype=mime)
    encoded_name = quote(str(filename))
    response.headers["Content-Disposition"] = f"inline; filename*=UTF-8''{encoded_name}"
    return response


@app.post("/api/positions/<position_id>/attachment")
def api_upload_position_attachment(position_id):
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"error": "Wybierz plik do zalaczenia."}), 400

    filename = str(upload.filename or "").strip()
    mime = str(upload.mimetype or "application/octet-stream").strip() or "application/octet-stream"
    payload = upload.read()
    if payload is None:
        payload = b""
    if len(payload) <= 0:
        return jsonify({"error": "Wybrany plik jest pusty."}), 400
    if len(payload) > MAX_ATTACHMENT_BYTES:
        return jsonify({"error": f"Plik jest za duzy. Maksymalny rozmiar to {MAX_ATTACHMENT_BYTES // (1024 * 1024)} MB."}), 413

    attachment_data = base64.b64encode(payload).decode("ascii")
    connection = db()
    row = connection.execute("SELECT id, order_id FROM order_positions WHERE id = ?", (position_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pozycji."}), 404

    connection.execute(
        "UPDATE order_positions SET attachment_name = ?, attachment_mime = ?, attachment_data = ? WHERE id = ?",
        (filename, mime, attachment_data, position_id),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.delete("/api/positions/<position_id>/attachment")
def api_delete_position_attachment(position_id):
    connection = db()
    row = connection.execute("SELECT id FROM order_positions WHERE id = ?", (position_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pozycji."}), 404
    connection.execute(
        "UPDATE order_positions SET attachment_name = NULL, attachment_mime = NULL, attachment_data = NULL WHERE id = ?",
        (position_id,),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.post("/api/positions/bulk-status")
def api_bulk_position_status():
    data = request.get_json(force=True)
    order_id = data.get("orderId")
    position_ids = data.get("positionIds", [])
    action = data.get("action")
    actor = data.get("actor", "Brygadzista")
    if not order_id or not isinstance(position_ids, list) or not position_ids or action not in ("start", "finish"):
        return jsonify({"error": "NieprawidĹ‚owe dane statusu pozycji."}), 400

    connection = db()
    placeholders = ",".join("?" for _ in position_ids)
    rows = connection.execute(
        f"SELECT id, status, started_at FROM order_positions WHERE id IN ({placeholders})",
        position_ids,
    ).fetchall()

    timestamp = now_iso()
    for row in rows:
        if action == "start" and row["status"] == "pending":
            connection.execute(
                """
                UPDATE order_positions
                SET status = ?, started_at = ?,
                    progress_percent = CASE
                      WHEN progress_percent IS NULL OR progress_percent <= 0 THEN 1
                      ELSE progress_percent
                    END,
                    current_department_status = CASE
                      WHEN current_department_status = 'Dokumentacja' THEN 'Maszynownia'
                      ELSE current_department_status
                    END
                WHERE id = ?
                """,
                ("in_progress", timestamp, row["id"]),
            )
            connection.execute(
                "INSERT INTO feedback_events (id, order_id, position_id, action, actor, at) VALUES (?, ?, ?, ?, ?, ?)",
                (str(uuid.uuid4()), order_id, row["id"], "start", actor, timestamp),
            )
        if action == "finish" and row["status"] != "done":
            started_at = row["started_at"] or timestamp
            connection.execute(
                "UPDATE order_positions SET status = ?, started_at = ?, finished_at = ?, current_department_status = ?, progress_percent = ? WHERE id = ?",
                ("done", started_at, timestamp, "Zakonczone", 100.0, row["id"]),
            )
            connection.execute(
                "INSERT INTO feedback_events (id, order_id, position_id, action, actor, at) VALUES (?, ?, ?, ?, ?, ?)",
                (str(uuid.uuid4()), order_id, row["id"], "finish", actor, timestamp),
            )

    sync_order_status_from_positions(connection, order_id)
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.post("/api/auth/login")
def api_auth_login():
    data = request.get_json(force=True)
    login = str(data.get("login", "")).strip()
    password = str(data.get("password", ""))
    database_key = str(data.get("databaseKey", "")).strip() or get_request_database_key()
    if not login or not password:
        return jsonify({"error": "Podaj login i haslo."}), 400
    target_path = database_key_to_path(database_key)
    if not target_path:
        return jsonify({"error": "Nieprawidlowy identyfikator bazy."}), 400
    if not target_path.exists():
        return jsonify({"error": "Wybrana baza nie istnieje."}), 404
    initialize(target_path)

    target_row = get_user_auth_row_from_path(target_path, login)
    base_row = None
    if Path(target_path) != BASE_DB_PATH:
        base_row = get_user_auth_row_from_path(BASE_DB_PATH, login)

    authenticated_row = None
    authenticated_source = None
    for source_path, row in ((target_path, target_row), (BASE_DB_PATH, base_row)):
        if not row:
            continue
        if not verify_password(password, row["password_hash"]):
            continue
        if not bool(row["active"]):
            return jsonify({"error": "Uzytkownik jest nieaktywny."}), 403
        authenticated_row = row
        authenticated_source = Path(source_path)
        break

    if not authenticated_row:
        return jsonify({"error": "Nieprawidlowy login lub haslo."}), 401

    user_payload = user_payload_from_row(authenticated_row)
    access_map = load_database_access_map()
    if not can_user_access_database(user_payload, database_key, access_map):
        return jsonify({"error": "Brak dostepu do wybranej bazy."}), 403

    if authenticated_source and authenticated_source != Path(target_path):
        upsert_user_row_to_database(target_path, authenticated_row)
        refreshed = get_user_auth_row_from_path(target_path, login)
        if refreshed:
            user_payload = user_payload_from_row(refreshed)

    session.clear()
    set_request_database_key(database_key)
    session["user_id"] = user_payload["id"]
    g.current_user = user_payload
    return jsonify(
        {
            "ok": True,
            "user": user_payload,
            "activeDatabase": get_request_database_key(),
            "databases": list_database_catalog_for_user(user_payload, get_request_database_key(), access_map),
        }
    )


@app.get("/api/auth/session")
def api_auth_session():
    user = current_authenticated_user()
    if not user:
        return jsonify({"error": "Brak aktywnej sesji."}), 401
    selected_key = get_request_database_key()
    access_map = load_database_access_map()
    if not can_user_access_database(user, selected_key, access_map):
        session.pop("user_id", None)
        g.current_user = None
        return jsonify({"error": "Brak dostepu do aktualnej bazy. Zaloguj sie ponownie."}), 403
    return jsonify(
        {
            "ok": True,
            "user": user,
            "activeDatabase": selected_key,
            "databases": list_database_catalog_for_user(user, selected_key, access_map),
        }
    )


@app.post("/api/auth/logout")
def api_auth_logout():
    selected_db = get_request_database_key()
    session.clear()
    set_request_database_key(selected_db)
    g.current_user = None
    return jsonify({"ok": True})


def require_admin():
    user = current_authenticated_user()
    role = str((user or {}).get("role", "")).strip().lower()
    if role != "admin":
        return jsonify({"error": "Tylko administrator moze wykonac te operacje."}), 403
    return None


def can_user_create_databases(user):
    role = str((user or {}).get("role", "")).strip().lower()
    if role == "admin":
        return True
    return bool((user or {}).get("canCreateDatabases"))


def sync_session_after_database_switch(preferred_login, database_key=None):
    login = str(preferred_login or "").strip()
    key = str(database_key or "").strip() or get_request_database_key()
    target_path = database_key_to_path(key)
    if not target_path or not target_path.exists():
        session.pop("user_id", None)
        g.current_user = None
        return False
    if not login:
        session.pop("user_id", None)
        g.current_user = None
        return False
    row = get_user_auth_row_from_path(target_path, login)
    if (not row or not bool(row["active"])) and target_path != BASE_DB_PATH:
        fallback = get_user_auth_row_from_path(BASE_DB_PATH, login)
        if fallback and bool(fallback["active"]):
            fallback_payload = user_payload_from_row(fallback)
            if can_user_access_database(fallback_payload, key):
                upsert_user_row_to_database(target_path, fallback)
                row = get_user_auth_row_from_path(target_path, login) or fallback
    if not row:
        session.pop("user_id", None)
        g.current_user = None
        return False
    if not bool(row["active"]):
        session.pop("user_id", None)
        g.current_user = None
        return False
    user_payload = user_payload_from_row(row)
    set_request_database_key(key)
    session["user_id"] = user_payload["id"]
    g.current_user = user_payload
    return True


@app.get("/api/public/databases")
def api_public_databases():
    selected_key = get_request_database_key()
    return jsonify(
        {
            "ok": True,
            "activeDatabase": selected_key,
            "databases": list_database_catalog(selected_key),
        }
    )


@app.get("/api/databases")
def api_get_databases():
    user = current_authenticated_user()
    selected_key = get_request_database_key()
    return jsonify({"databases": list_database_catalog_for_user(user, selected_key), "activeDatabase": selected_key})


@app.get("/api/database-access")
def api_get_database_access():
    deny = require_admin_user_management()
    if deny:
        return deny
    selected_key = get_request_database_key()
    return jsonify(
        {
            "ok": True,
            "databaseAccessMap": load_database_access_map(),
            "databases": list_database_catalog(selected_key),
            "activeDatabase": selected_key,
        }
    )


@app.put("/api/database-access")
def api_update_database_access():
    deny = require_admin_user_management()
    if deny:
        return deny
    data = request.get_json(force=True)
    raw_map = data.get("databaseAccessMap")
    if raw_map is None:
        raw_map = data.get("map", {})
    normalized = save_database_access_map(raw_map)
    selected_key = get_request_database_key()
    return jsonify(
        {
            "ok": True,
            "databaseAccessMap": normalized,
            "databases": list_database_catalog(selected_key),
            "activeDatabase": selected_key,
        }
    )


@app.post("/api/databases")
def api_create_database():
    user = current_authenticated_user()
    if not can_user_create_databases(user):
        return jsonify({"error": "Brak uprawnien do tworzenia baz."}), 403
    data = request.get_json(force=True)
    name = str(data.get("name", "")).strip()
    slug = sanitize_variant_slug(name)
    if not slug:
        return jsonify({"error": "Podaj nazwe nowej bazy (litery/cyfry/-/_)."}), 400
    database_key = f"variant:{slug}"
    target_path = database_key_to_path(database_key)
    if not target_path:
        return jsonify({"error": "Nieprawidlowa nazwa bazy."}), 400
    if target_path.exists():
        return jsonify({"error": "Baza o tej nazwie juz istnieje."}), 409

    creator_login = str((user or {}).get("login", "")).strip()
    source_connection = db()
    initialize(target_path)
    target_connection = db_connect(target_path)
    source_users = source_connection.execute(
        """
        SELECT id, name, department, login, password_hash, role, permissions_json, active, can_create_databases
        FROM users
        WHERE COALESCE(login, '') <> ''
        """
    ).fetchall()
    source_skill_workers = source_connection.execute(
        """
        SELECT id, name, department, primary_station_id, active
        FROM skill_workers
        """
    ).fetchall()
    source_skill_levels = source_connection.execute(
        """
        SELECT worker_id, station_id, level
        FROM skill_worker_skills
        """
    ).fetchall()
    source_skill_availability = source_connection.execute(
        """
        SELECT worker_id, day, shift, minutes
        FROM skill_worker_availability
        """
    ).fetchall()
    if source_users:
        target_connection.execute("DELETE FROM users")
        target_connection.executemany(
            """
            INSERT INTO users (
              id, name, department, login, password_hash, role, permissions_json, active, can_create_databases
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    str(row["id"] or uuid.uuid4()),
                    str(row["name"] or "").strip() or "Uzytkownik",
                    normalize_standard_department(row["department"]),
                    str(row["login"] or "").strip(),
                    str(row["password_hash"] or "").strip(),
                    str(row["role"] or "user").strip().lower() if str(row["role"] or "").strip() else "user",
                    str(row["permissions_json"] or "[]"),
                    1 if bool(row["active"]) else 0,
                    1 if bool(row["can_create_databases"]) else 0,
                )
                for row in source_users
            ],
        )
    if source_skill_workers:
        target_connection.execute("DELETE FROM skill_worker_skills")
        target_connection.execute("DELETE FROM skill_workers")
        target_connection.executemany(
            "INSERT INTO skill_workers (id, name, department, primary_station_id, active) VALUES (?, ?, ?, ?, ?)",
            [
                (
                    str(row["id"] or uuid.uuid4()),
                    str(row["name"] or "").strip() or "Pracownik",
                    normalize_skill_department(row["department"]),
                    str(row["primary_station_id"] or "").strip() or None,
                    1 if bool(row["active"]) else 0,
                )
                for row in source_skill_workers
            ],
        )
        target_connection.execute(
            """
            UPDATE skill_workers
            SET primary_station_id = NULL
            WHERE COALESCE(primary_station_id, '') <> ''
              AND primary_station_id NOT IN (SELECT id FROM stations)
            """
        )
    if source_skill_levels:
        worker_ids = {
            str(row["id"] or "")
            for row in target_connection.execute("SELECT id FROM skill_workers").fetchall()
        }
        station_ids = {
            str(row["id"] or "")
            for row in target_connection.execute("SELECT id FROM stations").fetchall()
        }
        rows_to_insert = []
        for row in source_skill_levels:
            worker_id = str(row["worker_id"] or "")
            station_id = str(row["station_id"] or "")
            if worker_id not in worker_ids or station_id not in station_ids:
                continue
            level = clamp_int(row["level"], 0, 3, 0)
            if level <= 0:
                continue
            rows_to_insert.append((worker_id, station_id, level))
        if rows_to_insert:
            target_connection.executemany(
                "INSERT INTO skill_worker_skills (worker_id, station_id, level) VALUES (?, ?, ?)",
                rows_to_insert,
            )
    if source_skill_availability:
        worker_ids = {
            str(row["id"] or "")
            for row in target_connection.execute("SELECT id FROM skill_workers").fetchall()
        }
        rows_to_insert = []
        for row in source_skill_availability:
            worker_id = str(row["worker_id"] or "").strip()
            day = str(row["day"] or "").strip()
            shift = clamp_int(row["shift"], 1, 3, 1)
            minutes = clamp_int(row["minutes"], 0, 1440, 0)
            if worker_id not in worker_ids or not is_valid_iso_date(day):
                continue
            rows_to_insert.append((worker_id, day, shift, minutes))
        if rows_to_insert:
            target_connection.executemany(
                """
                INSERT INTO skill_worker_availability (worker_id, day, shift, minutes)
                VALUES (?, ?, ?, ?)
                """,
                rows_to_insert,
            )
    if creator_login:
        creator_row = source_connection.execute(
            """
            SELECT id, name, department, login, password_hash, role, permissions_json, active, can_create_databases
            FROM users
            WHERE LOWER(COALESCE(login, '')) = LOWER(?)
            LIMIT 1
            """,
            (creator_login,),
        ).fetchone()
        if creator_row:
            existing_creator = target_connection.execute(
                "SELECT id FROM users WHERE LOWER(COALESCE(login, '')) = LOWER(?) LIMIT 1",
                (creator_login,),
            ).fetchone()
            creator_payload = (
                str(creator_row["name"] or "").strip() or "Uzytkownik",
                normalize_standard_department(creator_row["department"]),
                str(creator_row["password_hash"] or "").strip() or hash_password("admin"),
                str(creator_row["role"] or "user").strip().lower() if str(creator_row["role"] or "").strip() else "user",
                str(creator_row["permissions_json"] or "[]"),
                1,
                1 if bool(creator_row["can_create_databases"]) else 0,
            )
            if existing_creator:
                target_connection.execute(
                    """
                    UPDATE users
                    SET name = ?, department = ?, password_hash = ?, role = ?, permissions_json = ?, active = ?, can_create_databases = ?
                    WHERE id = ?
                    """,
                    (
                        creator_payload[0],
                        creator_payload[1],
                        creator_payload[2],
                        creator_payload[3],
                        creator_payload[4],
                        creator_payload[5],
                        creator_payload[6],
                        existing_creator["id"],
                    ),
                )
            else:
                target_connection.execute(
                    """
                    INSERT INTO users (
                      id, name, department, login, password_hash, role, permissions_json, active, can_create_databases
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(creator_row["id"] or uuid.uuid4()),
                        creator_payload[0],
                        creator_payload[1],
                        creator_login,
                        creator_payload[2],
                        creator_payload[3],
                        creator_payload[4],
                        creator_payload[5],
                        creator_payload[6],
                    ),
                )
    target_connection.execute("UPDATE users SET can_create_databases = 1 WHERE LOWER(COALESCE(role, '')) = 'admin'")
    target_connection.commit()
    target_connection.close()
    source_connection.close()

    activate = to_bool(data.get("activate"))
    relogin_required = False
    if activate:
        current_login = str((current_authenticated_user() or {}).get("login", "")).strip()
        set_request_database_key(database_key)
        relogin_required = not sync_session_after_database_switch(current_login, database_key)

    return jsonify(
        {
            "ok": True,
            "createdDatabase": database_key,
            "activeDatabase": get_request_database_key(),
            "databases": list_database_catalog_for_user(
                current_authenticated_user(),
                get_request_database_key(),
            ),
            "requiresRelogin": relogin_required,
        }
    )


@app.put("/api/databases/active")
def api_set_active_database():
    data = request.get_json(force=True)
    database_key = str(data.get("key", "")).strip()
    target_path = database_key_to_path(database_key)
    if not target_path:
        return jsonify({"error": "Nieprawidlowy identyfikator bazy."}), 400
    if not target_path.exists():
        return jsonify({"error": "Wybrana baza nie istnieje."}), 404

    user = current_authenticated_user() or {}
    if not can_user_access_database(user, database_key):
        return jsonify({"error": "Brak dostepu do wybranej bazy."}), 403

    current_login = str((user or {}).get("login", "")).strip()
    initialize(target_path)
    set_request_database_key(database_key)
    relogin_required = not sync_session_after_database_switch(current_login, database_key)

    return jsonify(
        {
            "ok": True,
            "activeDatabase": get_request_database_key(),
            "databases": list_database_catalog_for_user(
                current_authenticated_user(),
                get_request_database_key(),
            ),
            "requiresRelogin": relogin_required,
        }
    )


def require_admin_user_management():
    deny = require_admin()
    if deny:
        return jsonify({"error": "Tylko administrator moze zarzadzac uzytkownikami."}), 403
    return None


def normalize_standard_department(value):
    allowed = [str(item or "").strip() for item in DEPARTMENTS if str(item or "").strip()]
    if not allowed:
        return "Maszynownia"
    department = str(value or "").strip()
    if department in allowed:
        return department
    return allowed[0]


def normalize_skill_department(value):
    return normalize_standard_department(value)


def normalize_skills_payload(raw_skills, valid_station_ids):
    if not isinstance(raw_skills, dict):
        return {}
    output = {}
    for raw_station_id, raw_level in raw_skills.items():
        station_id = str(raw_station_id or "").strip()
        if not station_id or station_id not in valid_station_ids:
            continue
        level = clamp_int(raw_level, 0, 3, 0)
        if level <= 0:
            continue
        output[station_id] = level
    return output


def upsert_skill_worker(connection, worker_id, payload):
    worker_name = str(payload.get("name", "")).strip()
    if not worker_name:
        return "Imie i nazwisko pracownika jest wymagane.", 400
    department = normalize_skill_department(payload.get("department"))
    active = to_bool(payload.get("active", True))
    stations = get_stations(connection)
    station_by_id = {str(item.get("id") or ""): item for item in stations}
    valid_station_ids = set(station_by_id.keys())
    primary_station_id = str(payload.get("primaryStationId", "")).strip()
    if primary_station_id:
        station = station_by_id.get(primary_station_id)
        if not station:
            return "Wybrane stanowisko glowne nie istnieje.", 400
        if str(station.get("department") or "").strip() != department:
            return "Stanowisko glowne musi nalezec do tego samego dzialu bazowego.", 400
        if not bool(station.get("active", True)):
            return "Stanowisko glowne musi byc aktywne.", 400
    normalized_skills = normalize_skills_payload(payload.get("skills"), valid_station_ids)

    existing = connection.execute("SELECT id FROM skill_workers WHERE id = ?", (worker_id,)).fetchone()
    if existing:
        connection.execute(
            "UPDATE skill_workers SET name = ?, department = ?, primary_station_id = ?, active = ? WHERE id = ?",
            (worker_name, department, primary_station_id or None, 1 if active else 0, worker_id),
        )
    else:
        connection.execute(
            "INSERT INTO skill_workers (id, name, department, primary_station_id, active) VALUES (?, ?, ?, ?, ?)",
            (worker_id, worker_name, department, primary_station_id or None, 1 if active else 0),
        )
    connection.execute("DELETE FROM skill_worker_skills WHERE worker_id = ?", (worker_id,))
    if normalized_skills:
        connection.executemany(
            "INSERT INTO skill_worker_skills (worker_id, station_id, level) VALUES (?, ?, ?)",
            [(worker_id, station_id, level) for station_id, level in normalized_skills.items()],
        )
    return None, None


@app.post("/api/users")
def api_create_user():
    deny = require_admin_user_management()
    if deny:
        return deny
    data = request.get_json(force=True)
    name = str(data.get("name", "")).strip()
    if not name:
        return jsonify({"error": "ImiÄ™ i nazwisko jest wymagane."}), 400
    login = str(data.get("login", "")).strip()
    password = str(data.get("password", ""))
    if not login or not password:
        return jsonify({"error": "Login i haslo sa wymagane."}), 400
    department = normalize_standard_department(data.get("department"))
    role = str(data.get("role", "user")).strip().lower()
    role = "admin" if role == "admin" else "user"
    can_create_databases = to_bool(data.get("canCreateDatabases"))
    visible_sections = normalize_visible_sections(data.get("visibleSections"))
    if role == "admin":
        visible_sections = USER_VISIBLE_SECTIONS[:]
        can_create_databases = True
    elif len(visible_sections) == 0:
        return jsonify({"error": "Wybierz przynajmniej jedna sekcje widoczna dla uzytkownika."}), 400
    user_id = str(uuid.uuid4())
    connection = db()
    exists = connection.execute("SELECT id FROM users WHERE LOWER(COALESCE(login, '')) = LOWER(?)", (login,)).fetchone()
    if exists:
        connection.close()
        return jsonify({"error": "Login jest juz zajety."}), 409
    connection.execute(
        """
        INSERT INTO users (id, name, department, login, password_hash, role, permissions_json, active, can_create_databases)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            name,
            department,
            login,
            hash_password(password),
            role,
            json.dumps(visible_sections, ensure_ascii=True),
            1,
            1 if can_create_databases else 0,
        ),
    )
    connection.commit()
    connection.close()
    return jsonify({"id": user_id}), 201


@app.put("/api/users/<user_id>")
def api_update_user(user_id):
    deny = require_admin_user_management()
    if deny:
        return deny
    data = request.get_json(force=True)
    name = str(data.get("name", "")).strip()
    if not name:
        return jsonify({"error": "Imie i nazwisko jest wymagane."}), 400
    login = str(data.get("login", "")).strip()
    if not login:
        return jsonify({"error": "Login jest wymagany."}), 400
    password = str(data.get("password", ""))
    department = normalize_standard_department(data.get("department"))
    role = str(data.get("role", "user")).strip().lower()
    role = "admin" if role == "admin" else "user"
    can_create_databases = to_bool(data.get("canCreateDatabases"))
    visible_sections = normalize_visible_sections(data.get("visibleSections"))
    if role == "admin":
        visible_sections = USER_VISIBLE_SECTIONS[:]
        can_create_databases = True
    elif len(visible_sections) == 0:
        return jsonify({"error": "Wybierz przynajmniej jedna sekcje widoczna dla uzytkownika."}), 400

    connection = db()
    row = connection.execute("SELECT id, login FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono uzytkownika."}), 404
    previous_login = str(row["login"] or "").strip()

    exists = connection.execute(
        "SELECT id FROM users WHERE LOWER(COALESCE(login, '')) = LOWER(?) AND id <> ?",
        (login, user_id),
    ).fetchone()
    if exists:
        connection.close()
        return jsonify({"error": "Login jest juz zajety."}), 409

    if password:
        connection.execute(
            """
            UPDATE users
            SET name = ?, department = ?, login = ?, password_hash = ?, role = ?, permissions_json = ?, can_create_databases = ?
            WHERE id = ?
            """,
            (
                name,
                department,
                login,
                hash_password(password),
                role,
                json.dumps(visible_sections, ensure_ascii=True),
                1 if can_create_databases else 0,
                user_id,
            ),
        )
    else:
        connection.execute(
            """
            UPDATE users
            SET name = ?, department = ?, login = ?, role = ?, permissions_json = ?, can_create_databases = ?
            WHERE id = ?
            """,
            (
                name,
                department,
                login,
                role,
                json.dumps(visible_sections, ensure_ascii=True),
                1 if can_create_databases else 0,
                user_id,
            ),
        )
    connection.commit()
    connection.close()
    if normalize_login_key(previous_login) != normalize_login_key(login):
        rename_database_access_login(previous_login, login)
    return jsonify({"ok": True})


@app.delete("/api/users/<user_id>")
def api_delete_user(user_id):
    deny = require_admin_user_management()
    if deny:
        return deny
    connection = db()
    row = connection.execute("SELECT id, role, login FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono uzytkownika."}), 404

    role = str(row["role"] or "user").strip().lower()
    if role == "admin":
        connection.close()
        return jsonify({"error": "Konto administratora nie moze byc usuniete."}), 400

    connection.execute("DELETE FROM users WHERE id = ?", (user_id,))
    connection.commit()
    connection.close()
    remove_database_access_login(row["login"])
    return jsonify({"ok": True})


@app.post("/api/skills/workers")
def api_create_skill_worker():
    payload = request.get_json(force=True)
    worker_id = str(uuid.uuid4())
    connection = db()
    error_message, error_code = upsert_skill_worker(connection, worker_id, payload)
    if error_message:
        connection.close()
        return jsonify({"error": error_message}), error_code
    connection.commit()
    connection.close()
    return jsonify({"id": worker_id}), 201


@app.put("/api/skills/workers/<worker_id>")
def api_update_skill_worker(worker_id):
    payload = request.get_json(force=True)
    worker_id = str(worker_id or "").strip()
    if not worker_id:
        return jsonify({"error": "Brak identyfikatora pracownika."}), 400
    connection = db()
    row = connection.execute("SELECT id FROM skill_workers WHERE id = ?", (worker_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pracownika."}), 404
    error_message, error_code = upsert_skill_worker(connection, worker_id, payload)
    if error_message:
        connection.close()
        return jsonify({"error": error_message}), error_code
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.delete("/api/skills/workers/<worker_id>")
def api_delete_skill_worker(worker_id):
    worker_id = str(worker_id or "").strip()
    if not worker_id:
        return jsonify({"error": "Brak identyfikatora pracownika."}), 400
    connection = db()
    row = connection.execute("SELECT id FROM skill_workers WHERE id = ?", (worker_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pracownika."}), 404
    connection.execute("DELETE FROM skill_worker_skills WHERE worker_id = ?", (worker_id,))
    connection.execute("DELETE FROM skill_worker_availability WHERE worker_id = ?", (worker_id,))
    connection.execute("DELETE FROM skill_workers WHERE id = ?", (worker_id,))
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/skills/availability/bulk")
def api_upsert_skill_availability_bulk():
    payload = request.get_json(force=True)
    worker_id = str(payload.get("workerId", "")).strip()
    if not worker_id:
        return jsonify({"error": "Wybierz pracownika."}), 400

    entries = payload.get("entries")
    if not isinstance(entries, list):
        return jsonify({"error": "Nieprawidlowy format danych dostepnosci."}), 400

    start_date = normalize_optional_date(payload.get("startDate"))
    end_date = normalize_optional_date(payload.get("endDate"))
    if start_date and not is_valid_iso_date(start_date):
        return jsonify({"error": "Nieprawidlowa data startowa."}), 400
    if end_date and not is_valid_iso_date(end_date):
        return jsonify({"error": "Nieprawidlowa data koncowa."}), 400
    if start_date and end_date and end_date < start_date:
        return jsonify({"error": "Zakres dat jest nieprawidlowy."}), 400

    normalized = []
    for item in entries:
        if not isinstance(item, dict):
            continue
        date_value = str(item.get("date", "")).strip()
        if not is_valid_iso_date(date_value):
            continue
        shift = clamp_int(item.get("shift"), 1, 3, 1)
        minutes = clamp_int(item.get("minutes"), 0, 1440, 0)
        normalized.append((worker_id, date_value, shift, minutes))

    connection = db()
    row = connection.execute("SELECT id FROM skill_workers WHERE id = ?", (worker_id,)).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono pracownika."}), 404

    if start_date and end_date:
        connection.execute(
            "DELETE FROM skill_worker_availability WHERE worker_id = ? AND day >= ? AND day <= ?",
            (worker_id, start_date, end_date),
        )

    if normalized:
        connection.executemany(
            """
            INSERT INTO skill_worker_availability (worker_id, day, shift, minutes)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(worker_id, day, shift) DO UPDATE SET minutes = excluded.minutes
            """,
            normalized,
        )

    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/settings")
def api_update_settings():
    data = request.get_json(force=True)
    minutes_per_shift = int(data.get("minutesPerShift", 480))
    if minutes_per_shift < 60 or minutes_per_shift > 720:
        return jsonify({"error": "Minuty na zmianÄ™ muszÄ… byÄ‡ w zakresie 60-720."}), 400

    connection = db()
    current = get_settings(connection)
    if "workingDays" in data:
        working_days = normalize_working_days(data.get("workingDays"))
    else:
        working_days = current.get("working_days", [1, 2, 3, 4, 5])

    if "weekdayShifts" in data:
        weekday_shifts = normalize_weekday_shifts(data.get("weekdayShifts"), working_days)
    elif "workingDays" in data:
        weekday_shifts = normalize_weekday_shifts({}, working_days)
    else:
        weekday_shifts = normalize_weekday_shifts(current.get("weekday_shifts"), working_days)

    if not any(int(weekday_shifts.get(str(day), 0)) > 0 for day in range(7)):
        connection.close()
        return jsonify({"error": "Przynajmniej jeden dzien musi miec minimum 1 zmiane."}), 400
    working_days = [day for day in range(7) if int(weekday_shifts.get(str(day), 0)) > 0]

    if "calendarOverrides" in data:
        calendar_overrides = normalize_calendar_overrides(data.get("calendarOverrides"))
    else:
        calendar_overrides = current.get("calendar_overrides", {})
    connection.execute(
        """
        UPDATE settings
        SET minutes_per_shift = ?,
            working_days_json = ?,
            calendar_overrides_json = ?,
            weekday_shifts_json = ?
        WHERE id = 1
        """,
        (
            minutes_per_shift,
            json.dumps(working_days, ensure_ascii=True),
            json.dumps(calendar_overrides, ensure_ascii=True),
            json.dumps(weekday_shifts, ensure_ascii=True),
        ),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/settings/calendar-day")
def api_update_calendar_day():
    data = request.get_json(force=True)
    date_value = str(data.get("date", "")).strip()
    if not is_valid_iso_date(date_value):
        return jsonify({"error": "Nieprawidlowa data. Oczekiwano formatu RRRR-MM-DD."}), 400

    working = to_bool(data.get("working"))
    connection = db()
    current = get_settings(connection)
    overrides = dict(current.get("calendar_overrides", {}))
    overrides[date_value] = working
    connection.execute(
        "UPDATE settings SET calendar_overrides_json = ? WHERE id = 1",
        (json.dumps(overrides, ensure_ascii=True),),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/settings/station-overtime")
def api_upsert_station_overtime():
    data = request.get_json(force=True)
    date_value = str(data.get("date", "")).strip()
    station_id = str(data.get("stationId", "")).strip()
    minutes = clamp_int(data.get("minutes"), 0, 1440, 0)

    if not is_valid_iso_date(date_value):
        return jsonify({"error": "Nieprawidlowa data. Oczekiwano formatu RRRR-MM-DD."}), 400
    if not station_id:
        return jsonify({"error": "Wybierz stanowisko."}), 400
    if minutes <= 0:
        return jsonify({"error": "Minuty nadgodzin musza byc wieksze od 0."}), 400

    connection = db()
    valid_station_ids = {item["id"] for item in get_stations(connection)}
    if station_id not in valid_station_ids:
        connection.close()
        return jsonify({"error": "Wybrane stanowisko nie istnieje."}), 400

    current = get_settings(connection)
    overtime = normalize_station_overtime(current.get("station_overtime"))
    day_map = dict(overtime.get(date_value, {}))
    day_map[station_id] = minutes
    overtime[date_value] = day_map

    connection.execute(
        "UPDATE settings SET station_overtime_json = ? WHERE id = 1",
        (json.dumps(overtime, ensure_ascii=True),),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.delete("/api/settings/station-overtime")
def api_delete_station_overtime():
    data = request.get_json(force=True)
    date_value = str(data.get("date", "")).strip()
    station_id = str(data.get("stationId", "")).strip()

    if not is_valid_iso_date(date_value):
        return jsonify({"error": "Nieprawidlowa data. Oczekiwano formatu RRRR-MM-DD."}), 400
    if not station_id:
        return jsonify({"error": "Brak stanowiska do usuniecia."}), 400

    connection = db()
    current = get_settings(connection)
    overtime = normalize_station_overtime(current.get("station_overtime"))
    day_map = dict(overtime.get(date_value, {}))
    if station_id in day_map:
        del day_map[station_id]
    if day_map:
        overtime[date_value] = day_map
    elif date_value in overtime:
        del overtime[date_value]

    connection.execute(
        "UPDATE settings SET station_overtime_json = ? WHERE id = 1",
        (json.dumps(overtime, ensure_ascii=True),),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/station-settings")
def api_update_station_settings():
    data = request.get_json(force=True)
    station_settings = data.get("stationSettings", {})
    if not isinstance(station_settings, dict):
        return jsonify({"error": "NieprawidĹ‚owy format ustawieĹ„ stanowisk."}), 400

    connection = db()
    valid_ids = {item["id"] for item in get_stations(connection)}
    for station_id, cfg in station_settings.items():
        if station_id not in valid_ids:
            continue
        shift_count = clamp_int(cfg.get("shiftCount"), 1, 3, 2)
        people_count = clamp_int(cfg.get("peopleCount"), 1, 200, 1)
        connection.execute(
            """
            INSERT INTO station_settings (station_id, shift_count, people_count)
            VALUES (?, ?, ?)
            ON CONFLICT(station_id) DO UPDATE SET shift_count = excluded.shift_count, people_count = excluded.people_count
            """,
            (station_id, shift_count, people_count),
        )

    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/stations-config")
def api_update_stations_config():
    data = request.get_json(force=True)
    stations_payload = data.get("stations", [])
    if not isinstance(stations_payload, list):
        return jsonify({"error": "Nieprawidlowy format listy stanowisk."}), 400

    connection = db()
    existing_rows = get_stations(connection)
    existing_ids = {item["id"] for item in existing_rows}
    normalized_rows = []
    used_ids = set()

    for raw in stations_payload:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name", "")).strip()
        if not name:
            connection.close()
            return jsonify({"error": "Nazwa stanowiska jest wymagana."}), 400
        department = str(raw.get("department", "")).strip()
        if department not in DEPARTMENTS:
            connection.close()
            return jsonify({"error": f"Nieprawidlowy dzial: {department}."}), 400

        station_id = str(raw.get("id", "")).strip().upper()
        if not station_id or station_id in used_ids:
            station_id = generate_station_id(existing_ids.union(used_ids), department)
        used_ids.add(station_id)

        normalized_rows.append(
            {
                "id": station_id,
                "name": name,
                "department": department,
                "active": to_bool(raw.get("active", True)),
                "shiftCount": clamp_int(raw.get("shiftCount"), 1, 3, 2),
                "peopleCount": clamp_int(raw.get("peopleCount"), 1, 200, 1),
                "sortOrder": len(normalized_rows) + 1,
            }
        )

    desired_ids = {item["id"] for item in normalized_rows}
    removed_ids = existing_ids - desired_ids

    for station_id in removed_ids:
        connection.execute("DELETE FROM station_settings WHERE station_id = ?", (station_id,))
        connection.execute("DELETE FROM skill_worker_skills WHERE station_id = ?", (station_id,))
        connection.execute("DELETE FROM stations WHERE id = ?", (station_id,))

    for item in normalized_rows:
        connection.execute(
            """
            INSERT INTO stations (id, name, department, active, sort_order)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
              name = excluded.name,
              department = excluded.department,
              active = excluded.active,
              sort_order = excluded.sort_order
            """,
            (
                item["id"],
                item["name"],
                item["department"],
                int(bool(item["active"])),
                item["sortOrder"],
            ),
        )
        connection.execute(
            """
            INSERT INTO station_settings (station_id, shift_count, people_count)
            VALUES (?, ?, ?)
            ON CONFLICT(station_id) DO UPDATE SET shift_count = excluded.shift_count, people_count = excluded.people_count
            """,
            (item["id"], item["shiftCount"], item["peopleCount"]),
        )

    sync_technology_allocations_with_stations(connection)
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.put("/api/material-rules")
def api_update_material_rules():
    data = request.get_json(force=True)
    material_rules = data.get("materialRules", {})
    if not isinstance(material_rules, dict):
        return jsonify({"error": "Nieprawidlowy format mapowania materialow."}), 400

    normalized = {dept: [] for dept in DEPARTMENTS}
    for dept, materials in material_rules.items():
        if dept not in normalized or not isinstance(materials, list):
            continue
        normalized[dept] = [item for item in materials if item in MATERIAL_KEYS]

    connection = db()
    connection.execute(
        "UPDATE material_rules SET data_json = ? WHERE id = 1",
        (json.dumps(normalized, ensure_ascii=True),),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


@app.post("/api/technologies")
def api_create_technology():
    data = request.get_json(force=True)
    name = str(data.get("name", "")).strip()
    if not name:
        return jsonify({"error": "Nazwa technologii jest wymagana."}), 400

    connection = db()
    exists = connection.execute(
        "SELECT technology_name FROM technology_allocations WHERE technology_name = ?",
        (name,),
    ).fetchone()
    if exists:
        connection.close()
        return jsonify({"error": "Technologia o tej nazwie juĹĽ istnieje."}), 409

    stations = get_stations(connection)
    machining_ids = [item["id"] for item in stations if item["department"] == "Maszynownia"]
    painting_ids = [item["id"] for item in stations if item["department"] == "Lakiernia"]
    assembly_ids = [item["id"] for item in stations if item["department"] == "Kompletacja"]
    payload = {
        "machining": make_distribution([], machining_ids),
        "painting": make_distribution([], painting_ids),
        "assembly": make_distribution([], assembly_ids),
    }
    connection.execute(
        "INSERT INTO technology_allocations (technology_name, data_json) VALUES (?, ?)",
        (name, json.dumps(payload, ensure_ascii=True)),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True}), 201


@app.put("/api/technologies/<technology_name>/process/<process_name>")
def api_update_technology_process(technology_name, process_name):
    if process_name not in ("machining", "painting", "assembly"):
        return jsonify({"error": "Nieznany proces."}), 400

    data = request.get_json(force=True)
    percentages = data.get("percentages", {})
    if not isinstance(percentages, dict):
        return jsonify({"error": "NieprawidĹ‚owy format procentĂłw."}), 400

    connection = db()
    row = connection.execute(
        "SELECT data_json FROM technology_allocations WHERE technology_name = ?",
        (technology_name,),
    ).fetchone()
    if not row:
        connection.close()
        return jsonify({"error": "Nie znaleziono technologii."}), 404

    full = json.loads(row["data_json"])
    cleaned = {}
    for key, value in percentages.items():
        cleaned[str(key)] = max(0.0, parse_float(value, 0.0))
    full[process_name] = cleaned

    connection.execute(
        "UPDATE technology_allocations SET data_json = ? WHERE technology_name = ?",
        (json.dumps(full, ensure_ascii=True), technology_name),
    )
    connection.commit()
    connection.close()
    return jsonify({"ok": True})


def generate_station_id(existing_ids, department):
    prefix_map = {
        "Maszynownia": "M",
        "Lakiernia": "P",
        "Kompletacja": "K",
        "Kosmetyka": "X",
    }
    prefix = prefix_map.get(department, "S")
    number = 1
    normalized = {str(item).upper() for item in existing_ids}
    while f"{prefix}{number}" in normalized:
        number += 1
    return f"{prefix}{number}"


def sync_technology_allocations_with_stations(connection):
    stations = get_stations(connection)
    by_process = {
        "machining": {item["id"] for item in stations if item["department"] == "Maszynownia"},
        "painting": {item["id"] for item in stations if item["department"] == "Lakiernia"},
        "assembly": {item["id"] for item in stations if item["department"] == "Kompletacja"},
    }
    rows = connection.execute("SELECT technology_name, data_json FROM technology_allocations").fetchall()
    for row in rows:
        payload = json.loads(row["data_json"])
        changed = False
        for process_name in ("machining", "painting", "assembly"):
            allowed_ids = by_process[process_name]
            current = payload.get(process_name, {})
            if not isinstance(current, dict):
                current = {}
            cleaned = {}
            for station_id, value in current.items():
                sid = str(station_id)
                if sid in allowed_ids:
                    cleaned[sid] = max(0.0, parse_float(value, 0.0))
            for station_id in allowed_ids:
                if station_id not in cleaned:
                    cleaned[station_id] = 0.0
            if cleaned != current:
                payload[process_name] = cleaned
                changed = True
        if changed:
            connection.execute(
                "UPDATE technology_allocations SET data_json = ? WHERE technology_name = ?",
                (json.dumps(payload, ensure_ascii=True), row["technology_name"]),
            )


def hash_password(password):
    return hashlib.sha256(str(password or "").encode("utf-8")).hexdigest()


def verify_password(raw_password, password_hash):
    if not password_hash:
        return False
    return hash_password(raw_password) == str(password_hash)


def normalize_visible_sections(value):
    if not isinstance(value, list):
        return []
    out = []
    for item in value:
        section = str(item or "").strip()
        if section in USER_VISIBLE_SECTIONS and section not in out:
            out.append(section)
    return out


def normalize_optional_date(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def normalize_optional_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def clamp_int(value, minimum, maximum, fallback):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return fallback
    return max(minimum, min(maximum, parsed))


def parse_float(value, fallback):
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def normalize_working_days(value):
    if not isinstance(value, list):
        return [1, 2, 3, 4, 5]
    cleaned = []
    for item in value:
        try:
            day = int(item)
        except (TypeError, ValueError):
            continue
        if 0 <= day <= 6 and day not in cleaned:
            cleaned.append(day)
    return cleaned or [1, 2, 3, 4, 5]


def normalize_weekday_shifts(value, fallback_working_days=None):
    base = {str(day): 0 for day in range(7)}
    fallback_days = normalize_working_days(fallback_working_days or [1, 2, 3, 4, 5])
    for day in fallback_days:
        base[str(day)] = 2

    if isinstance(value, dict):
        for raw_day, raw_shift in value.items():
            try:
                day = int(raw_day)
            except (TypeError, ValueError):
                continue
            if day < 0 or day > 6:
                continue
            base[str(day)] = clamp_int(raw_shift, 0, 3, base[str(day)])

    has_any_working = any(int(base[str(day)]) > 0 for day in range(7))
    return base if has_any_working else dict(DEFAULT_WEEKDAY_SHIFTS)


def normalize_calendar_overrides(value):
    if not isinstance(value, dict):
        return {}
    cleaned = {}
    for key, item in value.items():
        date_value = str(key).strip()
        if not is_valid_iso_date(date_value):
            continue
        cleaned[date_value] = to_bool(item)
    return cleaned


def normalize_station_overtime(value):
    if not isinstance(value, dict):
        return {}
    cleaned = {}
    for key, item in value.items():
        date_value = str(key).strip()
        if not is_valid_iso_date(date_value):
            continue
        if not isinstance(item, dict):
            continue
        day_map = {}
        for raw_station_id, raw_minutes in item.items():
            station_id = str(raw_station_id or "").strip()
            if not station_id:
                continue
            minutes = clamp_int(raw_minutes, 0, 1440, 0)
            if minutes > 0:
                day_map[station_id] = minutes
        if day_map:
            cleaned[date_value] = day_map
    return cleaned


def is_valid_iso_date(value):
    try:
        datetime.strptime(str(value), "%Y-%m-%d")
    except ValueError:
        return False
    return True


def to_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in ("1", "true", "tak", "yes", "on")


def initialize_runtime():
    ensure_database_storage()
    initialize(BASE_DB_PATH)
    active_path = get_current_database_path()
    if active_path != BASE_DB_PATH:
        initialize(active_path)


initialize_runtime()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8000")), debug=False)

