from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def write_headers(sheet, headers):
    fill = PatternFill(start_color="DDEAD0", end_color="DDEAD0", fill_type="solid")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
        sheet.column_dimensions[cell.column_letter].width = max(14, len(header) + 2)


def main():
    repo_root = Path(__file__).resolve().parents[1]
    target_dir = repo_root / "templates"
    target_dir.mkdir(parents=True, exist_ok=True)
    target_file = target_dir / "import_zamowien_wzorzec.xlsx"

    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Zamowienia"
    positions_sheet = workbook.create_sheet("Pozycje")

    order_headers = [
        "orderNumber",
        "entryDate",
        "owner",
        "client",
        "orderStatus",
        "color",
        "extras",
        "manualStartDate",
        "manualPlannedDate",
    ]
    write_headers(orders_sheet, order_headers)
    orders_sheet.append(
        [
            "WGT-2026-001",
            "2026-04-20",
            "Piotr Planista",
            "Klient Test",
            "Dokumentacja",
            "RAL 7016",
            "Priorytet",
            "",
            "",
        ]
    )

    position_headers = [
        "orderNumber",
        "positionNumber",
        "width",
        "height",
        "technology",
        "line",
        "machiningTime",
        "paintingTime",
        "assemblyTime",
        "framesCount",
        "sashesCount",
        "shapeRect",
        "shapeSkos",
        "shapeLuk",
        "slemieCount",
        "slupekStalyCount",
        "przymykCount",
        "niskiProgCount",
        "notes",
        "currentDepartmentStatus",
        "workflowStatus",
        "progressPercent",
        "materialWoodDate",
        "materialWoodToOrder",
        "materialCorpusDate",
        "materialCorpusToOrder",
        "materialGlassDate",
        "materialGlassToOrder",
        "materialHardwareDate",
        "materialHardwareToOrder",
        "materialAccessoriesDate",
        "materialAccessoriesToOrder",
    ]
    write_headers(positions_sheet, position_headers)
    positions_sheet.append(
        [
            "WGT-2026-001",
            "1",
            1200,
            1500,
            "IV78 Drewno",
            "L1",
            210,
            140,
            160,
            1,
            1,
            True,
            False,
            False,
            0,
            0,
            0,
            0,
            "Pozycja testowa",
            "Dokumentacja",
            "pending",
            0,
            "2026-04-25",
            False,
            "2026-04-26",
            False,
            "",
            False,
            "2026-04-27",
            False,
            "",
            False,
        ]
    )

    workbook.save(target_file)
    workbook.close()
    print(f"Saved template: {target_file}")


if __name__ == "__main__":
    main()
